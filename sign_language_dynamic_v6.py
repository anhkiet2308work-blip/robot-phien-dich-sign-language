"""
🤖 ROBOT PHIÊN DỊCH NGÔN NGỮ KÝ HIỆU 2 CHIỀU
Ứng dụng học máy Deep Learning
ULTIMATE BEAUTIFUL VERSION

Tính năng:
- Giao diện cực đẹp
- Bảo mật bằng password cho Thu thập & Train
- Xuất Excel
- Hiển thị % training với progress bar
- Animation robot học tập
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog, simpledialog
import cv2
import numpy as np
import pickle, json, os, time
from datetime import datetime
from PIL import Image, ImageTk, ImageDraw, ImageFont
import threading
from collections import deque
import difflib
from pathlib import Path

os.environ.setdefault('TF_CPP_MIN_LOG_LEVEL','2')
import tensorflow as tf
from tensorflow import keras

try:
    import speech_recognition as sr
except:
    sr = None

try:
    import pyttsx3
except:
    pyttsx3 = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except:
    openpyxl = None


class UltimateSignLanguageApp:
    SEQ_LEN = 30
    
    # PASSWORD - Thay đổi ở đây
    PASSWORD = "admin123"
    
    # Color scheme - Modern & Beautiful
    COLOR_PRIMARY = '#3b82f6'      # Blue
    COLOR_SUCCESS = '#10b981'      # Green
    COLOR_WARNING = '#f59e0b'      # Orange
    COLOR_DANGER = '#ef4444'       # Red
    COLOR_PURPLE = '#8b5cf6'       # Purple
    COLOR_BG = '#f0f4f8'           # Light background
    COLOR_CARD = '#ffffff'         # White cards
    COLOR_TEXT = '#1e293b'         # Dark text
    COLOR_TEXT_LIGHT = '#64748b'   # Light text
    COLOR_BORDER = '#cbd5e1'       # Border
    COLOR_ACCENT = '#06b6d4'       # Cyan accent
    
    def __init__(self, root):
        self.root = root
        self.root.title("🤖 Robot Phiên Dịch Ngôn Ngữ Ký Hiệu 2 Chiều - Deep Learning")
        self.root.geometry("1450x920")
        self.root.configure(bg=self.COLOR_BG)
        
        # State
        self.current_mode = None
        self.is_locked = False
        
        # Camera
        self.cap = None
        self.is_running = False
        self.frame_buffer = deque(maxlen=self.SEQ_LEN)
        self.raw_roi_buffer = deque(maxlen=self.SEQ_LEN)
        
        # Collection
        self.is_collecting = False
        self.current_label = ""
        
        # Data
        self.dataset = {}
        self.labels_order = []
        self.model = None
        
        self.dataset_pkl = Path("dataset_dynamic.pkl")
        self.frames_root = Path("frames")
        self.frames_root.mkdir(exist_ok=True)
        self.frames_index_path = Path("frames_index.json")
        self.frames_index = {}
        
        self.lower_skin = np.array([0, 20, 70], dtype=np.uint8)
        self.upper_skin = np.array([20, 255, 255], dtype=np.uint8)
        
        # Speech
        self.recognizer = sr.Recognizer() if sr else None
        self.stt_is_listening = False
        
        # TTS
        self.tts = None
        self.tts_lock = threading.Lock()
        if pyttsx3:
            try:
                self.tts = pyttsx3.init()
                self.tts.setProperty('rate', 175)
                self.tts.setProperty('volume', 1.0)
            except:
                self.tts = None
        
        # Training progress
        self.training_progress = 0
        self.training_status = ""
        
        # Playback
        self.playback_frames = []
        self.playback_idx = 0
        self.playback_running = False
        self.playback_label = None
        
        # Robot animation
        self.robot_frame_idx = 0
        self.robot_animating = False
        
        # Load data
        self.load_dataset()
        self.load_frames_index()
        
        # Build UI
        self.build_ui()
        
        print("\n✅ 🤖 ROBOT PHIÊN DỊCH READY!\n")
    
    def build_ui(self):
        # ===== TOP HEADER - SUPER BEAUTIFUL =====
        header = tk.Frame(self.root, bg=self.COLOR_PRIMARY, height=100)
        header.pack(fill=tk.X, side=tk.TOP)
        header.pack_propagate(False)
        
        # Logo/Icon
        icon_frame = tk.Frame(header, bg=self.COLOR_PRIMARY)
        icon_frame.pack(side=tk.LEFT, padx=30, pady=20)
        
        tk.Label(
            icon_frame,
            text="🤖",
            font=('Arial', 48),
            bg=self.COLOR_PRIMARY,
            fg='white'
        ).pack()
        
        # Title
        title_frame = tk.Frame(header, bg=self.COLOR_PRIMARY)
        title_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=15)
        
        tk.Label(
            title_frame,
            text="ROBOT PHIÊN DỊCH NGÔN NGỮ KÝ HIỆU 2 CHIỀU",
            font=('Arial', 22, 'bold'),
            bg=self.COLOR_PRIMARY,
            fg='white'
        ).pack(anchor=tk.W)
        
        tk.Label(
            title_frame,
            text="Ứng dụng học máy Deep Learning - LSTM Neural Network",
            font=('Arial', 11),
            bg=self.COLOR_PRIMARY,
            fg='#bfdbfe'
        ).pack(anchor=tk.W, pady=(5, 0))
        
        # Version badge
        version_frame = tk.Frame(header, bg=self.COLOR_PRIMARY)
        version_frame.pack(side=tk.RIGHT, padx=30)
        
        tk.Label(
            version_frame,
            text="v2.0",
            font=('Arial', 10, 'bold'),
            bg='#1e40af',
            fg='white',
            padx=12,
            pady=5
        ).pack()
        
        tk.Label(
            version_frame,
            text="🔒 Protected",
            font=('Arial', 9),
            bg=self.COLOR_PRIMARY,
            fg='#bfdbfe'
        ).pack(pady=(5, 0))
        
        # ===== MODE SELECTOR - BEAUTIFUL CARDS =====
        mode_container = tk.Frame(self.root, bg=self.COLOR_BG)
        mode_container.pack(fill=tk.X, pady=20, padx=25)
        
        tk.Label(
            mode_container,
            text="⚡ Chọn chế độ hoạt động:",
            font=('Arial', 13, 'bold'),
            bg=self.COLOR_BG,
            fg=self.COLOR_TEXT
        ).pack(side=tk.LEFT, padx=10)
        
        self.mode_buttons = {}
        modes = [
            ("collect", "✋ Thu thập", self.COLOR_SUCCESS, "🔒"),
            ("train", "🧠 Huấn luyện", self.COLOR_PURPLE, "🔒"),
            ("recognize", "🎯 Nhận diện", self.COLOR_PRIMARY, ""),
            ("speech", "🎙️ Giọng nói", self.COLOR_WARNING, "")
        ]
        
        for mode_id, text, color, lock in modes:
            btn_text = f"{text} {lock}" if lock else text
            btn = tk.Button(
                mode_container,
                text=btn_text,
                font=('Arial', 11, 'bold'),
                bg=color,
                fg='white',
                relief=tk.FLAT,
                padx=22,
                pady=12,
                cursor='hand2',
                command=lambda m=mode_id: self.show_mode(m)
            )
            btn.pack(side=tk.LEFT, padx=5)
            self.mode_buttons[mode_id] = btn
        
        # ===== MAIN CONTAINER =====
        main = tk.Frame(self.root, bg=self.COLOR_BG)
        main.pack(fill=tk.BOTH, expand=True, padx=25, pady=(0, 20))
        
        # Left: Display
        left = tk.Frame(main, bg=self.COLOR_CARD, relief=tk.SOLID, bd=1)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 12))
        
        # Display header
        disp_header = tk.Frame(left, bg=self.COLOR_CARD, height=50)
        disp_header.pack(fill=tk.X)
        disp_header.pack_propagate(False)
        
        tk.Label(
            disp_header,
            text="📺 Màn hình hiển thị",
            font=('Arial', 13, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT
        ).pack(side=tk.LEFT, padx=20, pady=15)
        
        self.lbl_mode_indicator = tk.Label(
            disp_header,
            text="",
            font=('Arial', 11, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_PRIMARY
        )
        self.lbl_mode_indicator.pack(side=tk.RIGHT, padx=20)
        
        # Separator
        tk.Frame(left, bg=self.COLOR_BORDER, height=1).pack(fill=tk.X)
        
        # Canvas
        canvas_container = tk.Frame(left, bg='black', relief=tk.SUNKEN, bd=2)
        canvas_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
        
        self.canvas = tk.Canvas(canvas_container, bg='#1a1a1a', highlightthickness=0)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        
        # Camera controls
        self.cam_controls = tk.Frame(left, bg=self.COLOR_CARD)
        self.cam_controls.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        cam_btn_frame = tk.Frame(self.cam_controls, bg=self.COLOR_CARD)
        cam_btn_frame.pack(side=tk.LEFT)
        
        self.btn_cam_on = tk.Button(
            cam_btn_frame,
            text="📷 Bật Camera",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_PRIMARY,
            fg='white',
            relief=tk.FLAT,
            padx=18,
            pady=9,
            cursor='hand2',
            command=self.start_camera
        )
        self.btn_cam_on.pack(side=tk.LEFT, padx=3)
        
        self.btn_cam_off = tk.Button(
            cam_btn_frame,
            text="⏹ Tắt",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_DANGER,
            fg='white',
            relief=tk.FLAT,
            padx=18,
            pady=9,
            cursor='hand2',
            command=self.stop_camera,
            state=tk.DISABLED
        )
        self.btn_cam_off.pack(side=tk.LEFT, padx=3)
        
        cam_status = tk.Frame(self.cam_controls, bg=self.COLOR_CARD)
        cam_status.pack(side=tk.RIGHT)
        
        self.lbl_cam_status = tk.Label(
            cam_status,
            text="⚪ Tắt",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT_LIGHT
        )
        self.lbl_cam_status.pack(side=tk.LEFT, padx=10)
        
        self.lbl_buffer = tk.Label(
            cam_status,
            text="Buffer: 0/30",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_WARNING
        )
        self.lbl_buffer.pack(side=tk.LEFT, padx=10)
        
        # Right: Control Panel
        right = tk.Frame(main, bg=self.COLOR_CARD, relief=tk.SOLID, bd=1, width=520)
        right.pack(side=tk.RIGHT, fill=tk.BOTH, padx=(12, 0))
        right.pack_propagate(False)
        
        # Control header
        ctrl_header = tk.Frame(right, bg=self.COLOR_CARD, height=50)
        ctrl_header.pack(fill=tk.X)
        ctrl_header.pack_propagate(False)
        
        self.lbl_panel_title = tk.Label(
            ctrl_header,
            text="⚙️ Bảng điều khiển",
            font=('Arial', 13, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT
        )
        self.lbl_panel_title.pack(pady=15)
        
        # Separator
        tk.Frame(right, bg=self.COLOR_BORDER, height=1).pack(fill=tk.X)
        
        # Panels container
        self.panels_container = tk.Frame(right, bg=self.COLOR_CARD)
        self.panels_container.pack(fill=tk.BOTH, expand=True)
        
        # Build all panels
        self.build_collect_panel()
        self.build_train_panel()
        self.build_recognize_panel()
        self.build_speech_panel()
        
        # Hide all initially
        self.collect_panel.pack_forget()
        self.train_panel.pack_forget()
        self.recognize_panel.pack_forget()
        self.speech_panel.pack_forget()
        
        # Show welcome screen
        self.show_welcome()
    
    def show_welcome(self):
        """Show welcome screen when no mode selected"""
        welcome = tk.Frame(self.panels_container, bg=self.COLOR_CARD)
        welcome.pack(fill=tk.BOTH, expand=True, pady=50)
        
        tk.Label(
            welcome,
            text="🤖",
            font=('Arial', 80),
            bg=self.COLOR_CARD
        ).pack(pady=20)
        
        tk.Label(
            welcome,
            text="Chào mừng đến với\nRobot Phiên Dịch!",
            font=('Arial', 16, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT,
            justify=tk.CENTER
        ).pack(pady=10)
        
        tk.Label(
            welcome,
            text="Chọn chế độ bên trên để bắt đầu",
            font=('Arial', 11),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT_LIGHT
        ).pack()
    
    def build_collect_panel(self):
        self.collect_panel = tk.Frame(self.panels_container, bg=self.COLOR_CARD)
        
        # Input section
        input_section = tk.Frame(self.collect_panel, bg='#f8fafc', relief=tk.SOLID, bd=1)
        input_section.pack(fill=tk.X, padx=20, pady=15)
        
        tk.Label(
            input_section,
            text="📝 Tên động tác:",
            font=('Arial', 10, 'bold'),
            bg='#f8fafc',
            fg=self.COLOR_TEXT
        ).pack(pady=(10, 5), padx=15, anchor=tk.W)
        
        self.entry_label = tk.Entry(
            input_section,
            font=('Arial', 11),
            bg='white',
            fg=self.COLOR_TEXT,
            relief=tk.SOLID,
            bd=1,
            insertbackground=self.COLOR_TEXT
        )
        self.entry_label.pack(fill=tk.X, padx=15, pady=(0, 10), ipady=7)
        
        # Buttons
        btn_frame = tk.Frame(self.collect_panel, bg=self.COLOR_CARD)
        btn_frame.pack(pady=12)
        
        self.btn_start_collect = tk.Button(
            btn_frame,
            text="🎬 Bắt đầu ghi",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_SUCCESS,
            fg='white',
            relief=tk.FLAT,
            padx=18,
            pady=10,
            cursor='hand2',
            command=self.start_collecting
        )
        self.btn_start_collect.pack(side=tk.LEFT, padx=4)
        
        self.btn_stop_collect = tk.Button(
            btn_frame,
            text="⏸ Dừng",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_WARNING,
            fg='white',
            relief=tk.FLAT,
            padx=18,
            pady=10,
            cursor='hand2',
            command=self.stop_collecting,
            state=tk.DISABLED
        )
        self.btn_stop_collect.pack(side=tk.LEFT, padx=4)
        
        # Stats
        stats_frame = tk.Frame(self.collect_panel, bg='#f0f9ff', relief=tk.SOLID, bd=1)
        stats_frame.pack(fill=tk.X, padx=20, pady=(0, 15))
        
        self.lbl_collect_stats = tk.Label(
            stats_frame,
            text="Sequences: 0 | Tổng: 0",
            font=('Arial', 10, 'bold'),
            bg='#f0f9ff',
            fg=self.COLOR_PRIMARY
        )
        self.lbl_collect_stats.pack(pady=10)
        
        # Sequences list
        tk.Label(
            self.collect_panel,
            text="📋 Danh sách sequences:",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT
        ).pack(pady=(5, 5), padx=20, anchor=tk.W)
        
        list_frame = tk.Frame(self.collect_panel, bg=self.COLOR_CARD)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 10))
        
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.seq_listbox = tk.Listbox(
            list_frame,
            font=('Consolas', 9),
            bg='#f8fafc',
            fg=self.COLOR_TEXT,
            relief=tk.SOLID,
            bd=1,
            selectmode=tk.SINGLE,
            yscrollcommand=scrollbar.set
        )
        self.seq_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.seq_listbox.yview)
        
        # Delete buttons
        del_btn_frame = tk.Frame(self.collect_panel, bg=self.COLOR_CARD)
        del_btn_frame.pack(fill=tk.X, padx=20, pady=(0, 15))
        
        tk.Button(
            del_btn_frame,
            text="🗑️ Xóa",
            font=('Arial', 9, 'bold'),
            bg=self.COLOR_DANGER,
            fg='white',
            relief=tk.FLAT,
            padx=14,
            pady=6,
            cursor='hand2',
            command=self.delete_selected_sequence
        ).pack(side=tk.LEFT, padx=3)
        
        tk.Button(
            del_btn_frame,
            text="🗑️ Xóa label",
            font=('Arial', 9, 'bold'),
            bg=self.COLOR_DANGER,
            fg='white',
            relief=tk.FLAT,
            padx=14,
            pady=6,
            cursor='hand2',
            command=self.delete_entire_label
        ).pack(side=tk.LEFT, padx=3)
    
    def build_train_panel(self):
        self.train_panel = tk.Frame(self.panels_container, bg=self.COLOR_CARD)
        
        # Stats card
        stats_card = tk.Frame(self.train_panel, bg='#f0f9ff', relief=tk.SOLID, bd=1)
        stats_card.pack(fill=tk.X, padx=20, pady=15)
        
        self.lbl_train_stats = tk.Label(
            stats_card,
            text="Dataset: 0 nhãn, 0 sequences\nMô hình: Chưa train",
            font=('Arial', 11, 'bold'),
            bg='#f0f9ff',
            fg=self.COLOR_TEXT,
            justify=tk.CENTER
        )
        self.lbl_train_stats.pack(pady=15)
        
        # Progress bar
        self.progress_frame = tk.Frame(self.train_panel, bg=self.COLOR_CARD)
        
        tk.Label(
            self.progress_frame,
            text="📊 Tiến độ huấn luyện:",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT
        ).pack(pady=(0, 5))
        
        self.train_progress = ttk.Progressbar(
            self.progress_frame,
            length=400,
            mode='determinate',
            style='Custom.Horizontal.TProgressbar'
        )
        self.train_progress.pack(pady=5)
        
        self.lbl_train_progress = tk.Label(
            self.progress_frame,
            text="0%",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_PRIMARY
        )
        self.lbl_train_progress.pack(pady=5)
        
        # Robot animation
        self.robot_canvas = tk.Canvas(
            self.progress_frame,
            width=100,
            height=100,
            bg=self.COLOR_CARD,
            highlightthickness=0
        )
        self.robot_canvas.pack(pady=10)
        
        # Buttons
        btn_frame = tk.Frame(self.train_panel, bg=self.COLOR_CARD)
        btn_frame.pack(pady=15)
        
        self.btn_train = tk.Button(
            btn_frame,
            text="🚀 Huấn luyện mô hình",
            font=('Arial', 11, 'bold'),
            bg=self.COLOR_PURPLE,
            fg='white',
            relief=tk.FLAT,
            padx=25,
            pady=13,
            cursor='hand2',
            command=self.train_model
        )
        self.btn_train.pack(pady=6)
        
        tk.Button(
            btn_frame,
            text="📊 Xuất Excel",
            font=('Arial', 11, 'bold'),
            bg=self.COLOR_ACCENT,
            fg='white',
            relief=tk.FLAT,
            padx=25,
            pady=13,
            cursor='hand2',
            command=self.export_to_excel
        ).pack(pady=6)
        
        tk.Button(
            btn_frame,
            text="🗑️ Xóa dataset",
            font=('Arial', 11, 'bold'),
            bg=self.COLOR_DANGER,
            fg='white',
            relief=tk.FLAT,
            padx=25,
            pady=13,
            cursor='hand2',
            command=self.clear_dataset
        ).pack(pady=6)
        
        # Statistics table
        tk.Label(
            self.train_panel,
            text="📊 Thống kê chi tiết:",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT
        ).pack(pady=(5, 8), padx=20, anchor=tk.W)
        
        tree_frame = tk.Frame(self.train_panel, bg=self.COLOR_CARD)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 15))
        
        tree_scroll = tk.Scrollbar(tree_frame)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        style = ttk.Style()
        style.configure("Custom.Treeview", 
                       background='#f8fafc',
                       foreground=self.COLOR_TEXT,
                       rowheight=25,
                       fieldbackground='#f8fafc')
        style.configure("Custom.Treeview.Heading",
                       background=self.COLOR_PRIMARY,
                       foreground='white',
                       font=('Arial', 9, 'bold'))
        
        self.stats_tree = ttk.Treeview(
            tree_frame,
            columns=('label', 'sequences', 'frames'),
            show='headings',
            height=6,
            yscrollcommand=tree_scroll.set,
            style="Custom.Treeview"
        )
        tree_scroll.config(command=self.stats_tree.yview)
        
        self.stats_tree.heading('label', text='Label')
        self.stats_tree.heading('sequences', text='Sequences')
        self.stats_tree.heading('frames', text='Frames')
        
        self.stats_tree.column('label', width=180, anchor=tk.W)
        self.stats_tree.column('sequences', width=90, anchor=tk.CENTER)
        self.stats_tree.column('frames', width=90, anchor=tk.CENTER)
        
        self.stats_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Configure progressbar style
        style.configure("Custom.Horizontal.TProgressbar",
                       troughcolor='#e2e8f0',
                       background=self.COLOR_SUCCESS,
                       thickness=20)
    
    def build_recognize_panel(self):
        self.recognize_panel = tk.Frame(self.panels_container, bg=self.COLOR_CARD)
        
        tk.Label(
            self.recognize_panel,
            text="🎯 Kết quả nhận diện:",
            font=('Arial', 11, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT
        ).pack(pady=15)
        
        # Result display
        result_frame = tk.Frame(
            self.recognize_panel,
            bg='#eff6ff',
            relief=tk.SOLID,
            bd=2
        )
        result_frame.pack(fill=tk.X, padx=25, pady=10)
        
        self.lbl_pred = tk.Label(
            result_frame,
            text="—",
            font=('Arial', 52, 'bold'),
            bg='#eff6ff',
            fg=self.COLOR_PRIMARY
        )
        self.lbl_pred.pack(pady=25)
        
        self.lbl_conf = tk.Label(
            result_frame,
            text="Chờ buffer đầy...",
            font=('Arial', 12),
            bg='#eff6ff',
            fg=self.COLOR_TEXT_LIGHT
        )
        self.lbl_conf.pack(pady=(0, 25))
        
        # Info
        info_frame = tk.Frame(self.recognize_panel, bg='#f0fdf4', relief=tk.SOLID, bd=1)
        info_frame.pack(fill=tk.X, padx=25, pady=20)
        
        tk.Label(
            info_frame,
            text="🔊 Text-to-Speech",
            font=('Arial', 10, 'bold'),
            bg='#f0fdf4',
            fg=self.COLOR_SUCCESS
        ).pack(anchor=tk.W, padx=15, pady=(12, 5))
        
        tk.Label(
            info_frame,
            text="Robot sẽ tự động đọc tên động tác\nmỗi khi nhận diện thành công",
            font=('Arial', 9),
            bg='#f0fdf4',
            fg=self.COLOR_TEXT_LIGHT,
            justify=tk.LEFT
        ).pack(anchor=tk.W, padx=15, pady=(0, 12))
    
    def build_speech_panel(self):
        self.speech_panel = tk.Frame(self.panels_container, bg=self.COLOR_CARD)
        
        # Mic status
        mic_frame = tk.Frame(self.speech_panel, bg='#fef3c7', relief=tk.SOLID, bd=1)
        mic_frame.pack(fill=tk.X, padx=20, pady=15)
        
        self.lbl_mic = tk.Label(
            mic_frame,
            text="🎤 Microphone: Sẵn sàng",
            font=('Arial', 11, 'bold'),
            bg='#fef3c7',
            fg=self.COLOR_TEXT
        )
        self.lbl_mic.pack(pady=12)
        
        # Buttons
        btn_frame = tk.Frame(self.speech_panel, bg=self.COLOR_CARD)
        btn_frame.pack(pady=12)
        
        self.btn_listen = tk.Button(
            btn_frame,
            text="🎧 Bắt đầu nghe",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_SUCCESS,
            fg='white',
            relief=tk.FLAT,
            padx=18,
            pady=10,
            cursor='hand2',
            command=self.start_listening
        )
        self.btn_listen.pack(side=tk.LEFT, padx=4)
        
        self.btn_stop_listen = tk.Button(
            btn_frame,
            text="🛑 Dừng",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_DANGER,
            fg='white',
            relief=tk.FLAT,
            padx=18,
            pady=10,
            cursor='hand2',
            command=self.stop_listening,
            state=tk.DISABLED
        )
        self.btn_stop_listen.pack(side=tk.LEFT, padx=4)
        
        # Text display
        tk.Label(
            self.speech_panel,
            text="📝 Văn bản nhận dạng:",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT
        ).pack(pady=(15, 5), padx=20, anchor=tk.W)
        
        text_frame = tk.Frame(self.speech_panel, bg=self.COLOR_CARD)
        text_frame.pack(fill=tk.X, padx=20, pady=5)
        
        self.txt_speech = tk.Text(
            text_frame,
            height=3,
            font=('Arial', 11),
            bg='#f8fafc',
            fg=self.COLOR_TEXT,
            relief=tk.SOLID,
            bd=1,
            wrap=tk.WORD
        )
        self.txt_speech.pack(fill=tk.X, ipady=4)
        
        # Match result
        match_frame = tk.Frame(self.speech_panel, bg='#f0f9ff', relief=tk.SOLID, bd=1)
        match_frame.pack(fill=tk.X, padx=20, pady=15)
        
        self.lbl_match = tk.Label(
            match_frame,
            text="Khớp: —",
            font=('Arial', 13, 'bold'),
            bg='#f0f9ff',
            fg=self.COLOR_PRIMARY
        )
        self.lbl_match.pack(pady=12)
        
        # Info
        info_frame = tk.Frame(self.speech_panel, bg='#fef2f2', relief=tk.SOLID, bd=1)
        info_frame.pack(fill=tk.X, padx=20, pady=(0, 15))
        
        tk.Label(
            info_frame,
            text="📹 Animation Playback",
            font=('Arial', 10, 'bold'),
            bg='#fef2f2',
            fg=self.COLOR_DANGER
        ).pack(anchor=tk.W, padx=15, pady=(12, 5))
        
        tk.Label(
            info_frame,
            text="Khi tìm thấy label khớp,\nrobot sẽ phát animation tự động",
            font=('Arial', 9),
            bg='#fef2f2',
            fg=self.COLOR_TEXT_LIGHT,
            justify=tk.LEFT
        ).pack(anchor=tk.W, padx=15, pady=(0, 12))
    
    # ===== PASSWORD CHECK =====
    
    def check_password(self, action_name):
        """Check password for protected actions"""
        password = simpledialog.askstring(
            "🔒 Xác thực",
            f"Nhập mật khẩu để truy cập {action_name}:",
            show='*'
        )
        
        if password == self.PASSWORD:
            return True
        elif password is not None:  # User didn't cancel
            messagebox.showerror("Lỗi", "Mật khẩu sai!")
        return False
    
    # ===== MODE MANAGEMENT =====
    
    def show_mode(self, mode):
        # Check password for protected modes
        if mode == "collect" and not self.check_password("Thu thập mẫu"):
            return
        if mode == "train" and not self.check_password("Huấn luyện"):
            return
        
        if self.is_locked:
            messagebox.showwarning("Đang hoạt động", "Vui lòng dừng hoạt động hiện tại!")
            return
        
        self.current_mode = mode
        
        # Hide all panels
        self.collect_panel.pack_forget()
        self.train_panel.pack_forget()
        self.recognize_panel.pack_forget()
        self.speech_panel.pack_forget()
        
        # Stop playback
        self.playback_running = False
        
        # Update UI
        mode_info = {
            "collect": ("✋ THU THẬP MẪU", "📋 Quản lý thu thập", self.COLOR_SUCCESS),
            "train": ("🧠 HUẤN LUYỆN", "🚀 Huấn luyện & Thống kê", self.COLOR_PURPLE),
            "recognize": ("🎯 NHẬN DIỆN", "🎯 Nhận diện động tác", self.COLOR_PRIMARY),
            "speech": ("🎙️ GIỌNG NÓI", "🎙️ Chuyển giọng nói", self.COLOR_WARNING)
        }
        
        indicator, title, color = mode_info[mode]
        self.lbl_mode_indicator.config(text=indicator, fg=color)
        self.lbl_panel_title.config(text=title)
        
        # Show panel
        if mode == "collect":
            self.collect_panel.pack(fill=tk.BOTH, expand=True)
            self.refresh_sequences_list()
            self.update_stats()
        elif mode == "train":
            self.train_panel.pack(fill=tk.BOTH, expand=True)
            self.progress_frame.pack_forget()  # Hide initially
            self.refresh_stats_table()
            self.update_stats()
        elif mode == "recognize":
            self.recognize_panel.pack(fill=tk.BOTH, expand=True)
        elif mode == "speech":
            self.speech_panel.pack(fill=tk.BOTH, expand=True)
    
    def lock_interface(self, lock=True):
        """Lock/unlock interface"""
        self.is_locked = lock
        state = tk.DISABLED if lock else tk.NORMAL
        
        for btn in self.mode_buttons.values():
            btn.config(state=state)
    
    # ===== CAMERA =====
    
    def start_camera(self):
        if self.is_running:
            return
        
        self.cap = cv2.VideoCapture(0)
        if not self.cap.isOpened():
            messagebox.showerror("Lỗi", "Không mở được camera!")
            return
        
        self.is_running = True
        self.btn_cam_on.config(state=tk.DISABLED)
        self.btn_cam_off.config(state=tk.NORMAL)
        self.lbl_cam_status.config(text="🟢 Đang chạy", fg=self.COLOR_SUCCESS)
        self.frame_buffer.clear()
        
        self.playback_running = False
        self.update_frame()
    
    def stop_camera(self):
        if not self.is_running:
            return
        
        self.is_running = False
        if self.cap:
            self.cap.release()
        
        self.btn_cam_on.config(state=tk.NORMAL)
        self.btn_cam_off.config(state=tk.DISABLED)
        self.lbl_cam_status.config(text="⚪ Tắt", fg=self.COLOR_TEXT_LIGHT)
        self.canvas.delete("all")
        self.frame_buffer.clear()
    
    def update_frame(self):
        if not self.is_running:
            return
        
        ret, frame = self.cap.read()
        if not ret:
            self.root.after(10, self.update_frame)
            return
        
        frame = cv2.flip(frame, 1)
        cv2.rectangle(frame, (100, 100), (540, 380), (0, 255, 0), 2)
        cv2.putText(frame, "Dong tac o day", (110, 90),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
        
        roi = frame[100:380, 100:540]
        hsv = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
        mask = cv2.inRange(hsv, self.lower_skin, self.upper_skin)
        
        kernel = np.ones((3,3), np.uint8)
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel, iterations=2)
        mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel, iterations=2)
        mask = cv2.GaussianBlur(mask, (5, 5), 100)
        
        features = self.image_to_features(mask)
        if features is not None:
            self.frame_buffer.append(features)
            self.raw_roi_buffer.append(roi.copy())
        
        buffer_size = len(self.frame_buffer)
        self.lbl_buffer.config(text=f"Buffer: {buffer_size}/30")
        
        if buffer_size == self.SEQ_LEN:
            if self.is_collecting and self.current_label:
                self.collect_sequence()
            elif self.current_mode == "recognize" and self.model:
                self.recognize_current()
        
        mask_3ch = cv2.cvtColor(mask, cv2.COLOR_GRAY2BGR)
        frame[100:380, 100:540] = cv2.addWeighted(roi, 0.7, mask_3ch, 0.3, 0)
        
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(frame_rgb)
        
        canvas_w = self.canvas.winfo_width() or 800
        canvas_h = self.canvas.winfo_height() or 600
        img = img.resize((canvas_w, canvas_h), Image.LANCZOS)
        
        photo = ImageTk.PhotoImage(img)
        self.canvas.create_image(0, 0, anchor=tk.NW, image=photo)
        self.canvas.image = photo
        
        self.root.after(10, self.update_frame)
    
    def image_to_features(self, mask):
        try:
            resized = cv2.resize(mask, (28, 28))
            return resized.flatten() / 255.0
        except:
            return None
    
    # ===== COLLECTION =====
    
    def start_collecting(self):
        label = self.entry_label.get().strip()
        if not label:
            messagebox.showwarning("Cảnh báo", "Vui lòng nhập tên động tác!")
            return
        
        if not self.is_running:
            messagebox.showwarning("Cảnh báo", "Vui lòng bật camera!")
            return
        
        self.current_label = label
        self.is_collecting = True
        
        if label not in self.dataset:
            self.dataset[label] = []
            if label not in self.labels_order:
                self.labels_order.append(label)
        
        self.frame_buffer.clear()
        self.raw_roi_buffer.clear()
        
        self.btn_start_collect.config(state=tk.DISABLED)
        self.btn_stop_collect.config(state=tk.NORMAL)
        self.lock_interface(True)
    
    def stop_collecting(self):
        self.is_collecting = False
        self.current_label = ""
        
        self.btn_start_collect.config(state=tk.NORMAL)
        self.btn_stop_collect.config(state=tk.DISABLED)
        self.lock_interface(False)
        
        self.save_dataset()
        self.save_frames_index()
        self.refresh_sequences_list()
        self.update_stats()
    
    def collect_sequence(self):
        sequence = np.array(list(self.frame_buffer))
        self.dataset[self.current_label].append(sequence)
        
        count = len(self.dataset[self.current_label])
        self.update_stats()
        
        self.save_frames_for_sequence(self.current_label, count - 1)
        self.frame_buffer.clear()
        self.raw_roi_buffer.clear()
    
    def save_frames_for_sequence(self, label, seq_idx):
        try:
            label_dir = self.frames_root / label
            label_dir.mkdir(exist_ok=True)
            
            seq_dir = label_dir / f"seq_{seq_idx:04d}"
            seq_dir.mkdir(exist_ok=True)
            
            for i, roi in enumerate(self.raw_roi_buffer):
                frame_path = seq_dir / f"frame_{i:03d}.jpg"
                cv2.imwrite(str(frame_path), roi)
            
            if label not in self.frames_index:
                self.frames_index[label] = []
            
            rel_path = f"{label}/seq_{seq_idx:04d}"
            self.frames_index[label].append(rel_path)
        except Exception as e:
            print(f"✗ Error: {e}")
    
    def refresh_sequences_list(self):
        self.seq_listbox.delete(0, tk.END)
        
        label = self.entry_label.get().strip()
        if not label or label not in self.dataset:
            return
        
        sequences = self.dataset[label]
        for i in range(len(sequences)):
            self.seq_listbox.insert(tk.END, f"  Sequence {i+1}")
    
    def delete_selected_sequence(self):
        label = self.entry_label.get().strip()
        if not label or label not in self.dataset:
            messagebox.showwarning("Cảnh báo", "Chọn label!")
            return
        
        selection = self.seq_listbox.curselection()
        if not selection:
            messagebox.showwarning("Cảnh báo", "Chọn sequence!")
            return
        
        idx = selection[0]
        
        if not messagebox.askyesno("Xác nhận", f"Xóa Sequence {idx+1}?"):
            return
        
        del self.dataset[label][idx]
        
        if label in self.frames_index and idx < len(self.frames_index[label]):
            seq_rel = self.frames_index[label][idx]
            seq_dir = self.frames_root / seq_rel
            try:
                if seq_dir.exists():
                    for p in seq_dir.glob("*.jpg"):
                        p.unlink()
                    seq_dir.rmdir()
            except:
                pass
            
            del self.frames_index[label][idx]
        
        if len(self.dataset[label]) == 0:
            del self.dataset[label]
            self.labels_order.remove(label)
            if label in self.frames_index:
                del self.frames_index[label]
        
        self.save_dataset()
        self.save_frames_index()
        self.refresh_sequences_list()
        self.update_stats()
    
    def delete_entire_label(self):
        label = self.entry_label.get().strip()
        if not label or label not in self.dataset:
            messagebox.showwarning("Cảnh báo", "Label không tồn tại!")
            return
        
        seq_count = len(self.dataset[label])
        if not messagebox.askyesno("Xác nhận", 
            f"Xóa toàn bộ '{label}' ({seq_count} seq)?"):
            return
        
        del self.dataset[label]
        self.labels_order.remove(label)
        
        if label in self.frames_index:
            label_dir = self.frames_root / label
            try:
                if label_dir.exists():
                    for p in label_dir.rglob("*"):
                        if p.is_file():
                            p.unlink()
                    for p in sorted(label_dir.glob("*"), reverse=True):
                        if p.is_dir():
                            try:
                                p.rmdir()
                            except:
                                pass
                    label_dir.rmdir()
            except:
                pass
            
            del self.frames_index[label]
        
        self.save_dataset()
        self.save_frames_index()
        self.refresh_sequences_list()
        self.update_stats()
        self.entry_label.delete(0, tk.END)
    
    # ===== TRAINING =====
    
    def train_model(self):
        labels = [lb for lb in self.labels_order if lb in self.dataset and len(self.dataset[lb]) > 0]
        if len(labels) < 2:
            messagebox.showwarning("Cảnh báo", "Cần ít nhất 2 nhãn!")
            return
        
        self.lock_interface(True)
        self.btn_train.config(state=tk.DISABLED, text="⏳ Đang train...")
        
        # Show progress
        self.progress_frame.pack(fill=tk.X, padx=20, pady=15, before=self.btn_train.master)
        self.train_progress['value'] = 0
        self.lbl_train_progress.config(text="0%")
        
        # Start robot animation
        self.robot_animating = True
        self.animate_robot()
        
        threading.Thread(target=self._train_thread, daemon=True).start()
    
    def animate_robot(self):
        """Animate robot while training"""
        if not self.robot_animating:
            self.robot_canvas.delete("all")
            return
        
        self.robot_canvas.delete("all")
        
        # Simple robot animation
        frame = self.robot_frame_idx % 4
        
        # Body
        self.robot_canvas.create_oval(25, 30, 75, 80, fill='#3b82f6', outline='#1e40af', width=2)
        
        # Eyes (blink effect)
        if frame == 3:
            self.robot_canvas.create_line(35, 45, 45, 45, fill='white', width=3)
            self.robot_canvas.create_line(55, 45, 65, 45, fill='white', width=3)
        else:
            self.robot_canvas.create_oval(35, 40, 45, 50, fill='white')
            self.robot_canvas.create_oval(55, 40, 65, 50, fill='white')
        
        # Antenna (bounce)
        antenna_y = 20 - (frame % 2) * 3
        self.robot_canvas.create_line(50, 30, 50, antenna_y, fill='#1e40af', width=2)
        self.robot_canvas.create_oval(47, antenna_y-5, 53, antenna_y+1, fill='#ef4444')
        
        # Book (learning animation)
        book_x = 30 + (frame * 3)
        self.robot_canvas.create_rectangle(book_x, 55, book_x+15, 65, fill='#f59e0b', outline='#d97706')
        
        self.robot_frame_idx += 1
        self.root.after(200, self.animate_robot)
    
    def _train_thread(self):
        try:
            labels = [lb for lb in self.labels_order if lb in self.dataset and len(self.dataset[lb]) > 0]
            xs, ys = [], []
            
            for idx, label in enumerate(labels):
                for seq in self.dataset[label]:
                    xs.append(seq)
                    one_hot = np.zeros(len(labels))
                    one_hot[idx] = 1
                    ys.append(one_hot)
            
            xs = np.array(xs)
            ys = np.array(ys)
            
            self.model = keras.Sequential([
                keras.layers.LSTM(64, return_sequences=True, input_shape=(self.SEQ_LEN, 784)),
                keras.layers.Dropout(0.3),
                keras.layers.LSTM(32),
                keras.layers.Dropout(0.2),
                keras.layers.Dense(32, activation='relu'),
                keras.layers.Dense(len(labels), activation='softmax')
            ])
            
            self.model.compile(
                optimizer='adam',
                loss='categorical_crossentropy',
                metrics=['accuracy']
            )
            
            # Custom callback for progress
            class ProgressCallback(keras.callbacks.Callback):
                def __init__(self, app):
                    self.app = app
                    
                def on_epoch_end(self, epoch, logs=None):
                    progress = int((epoch + 1) / 50 * 100)
                    self.app.root.after(0, lambda: self.app.update_training_progress(progress))
            
            history = self.model.fit(
                xs, ys,
                epochs=50,
                batch_size=16,
                validation_split=0.2,
                verbose=0,
                callbacks=[ProgressCallback(self)]
            )
            
            acc = history.history['val_accuracy'][-1]
            
            self.root.after(0, lambda: self._on_train_complete(len(labels), len(xs), acc))
            
        except Exception as e:
            print(f"✗ Error: {e}")
            self.root.after(0, lambda: messagebox.showerror("Lỗi", f"Training thất bại: {e}"))
            self.root.after(0, lambda: self._cleanup_training())
    
    def update_training_progress(self, progress):
        """Update training progress bar"""
        self.train_progress['value'] = progress
        self.lbl_train_progress.config(text=f"{progress}%")
    
    def _on_train_complete(self, n_labels, n_seqs, acc):
        """Callback when training completes"""
        self.robot_animating = False
        time.sleep(0.3)
        
        self.lock_interface(False)
        self.btn_train.config(state=tk.NORMAL, text="🚀 Huấn luyện mô hình")
        self.update_stats()
        self.refresh_stats_table()
        
        messagebox.showinfo("Thành công!", 
            f"🎉 Mô hình đã được huấn luyện!\n\n"
            f"📊 Labels: {n_labels}\n"
            f"📋 Sequences: {n_seqs}\n"
            f"✅ Accuracy: {acc*100:.1f}%")
    
    def _cleanup_training(self):
        """Cleanup after training error"""
        self.robot_animating = False
        self.lock_interface(False)
        self.btn_train.config(state=tk.NORMAL, text="🚀 Huấn luyện mô hình")
    
    def clear_dataset(self):
        if not messagebox.askyesno("Xác nhận", "Xóa TOÀN BỘ dataset?"):
            return
        
        self.dataset = {}
        self.labels_order = []
        self.model = None
        self.frames_index = {}
        
        try:
            if self.frames_root.exists():
                for p in self.frames_root.rglob("*"):
                    if p.is_file():
                        p.unlink()
                for p in sorted(self.frames_root.glob("*"), reverse=True):
                    if p.is_dir():
                        try:
                            p.rmdir()
                        except:
                            pass
        except:
            pass
        
        self.save_dataset()
        self.save_frames_index()
        self.update_stats()
        self.refresh_stats_table()
    
    def refresh_stats_table(self):
        """Refresh statistics table"""
        for item in self.stats_tree.get_children():
            self.stats_tree.delete(item)
        
        for label in self.labels_order:
            if label in self.dataset:
                seq_count = len(self.dataset[label])
                frame_count = seq_count * self.SEQ_LEN
                self.stats_tree.insert('', tk.END, values=(label, seq_count, frame_count))
    
    def export_to_excel(self):
        """Export statistics to Excel"""
        if not openpyxl:
            messagebox.showerror("Lỗi", "Vui lòng cài: pip install openpyxl")
            return
        
        if len(self.labels_order) == 0:
            messagebox.showwarning("Cảnh báo", "Không có dữ liệu!")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"sign_language_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Statistics"
            
            # Header
            headers = ['STT', 'Label', 'Sequences', 'Frames', 'Total Samples']
            ws.append(headers)
            
            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data
            for idx, label in enumerate(self.labels_order, 1):
                if label in self.dataset:
                    seq_count = len(self.dataset[label])
                    frame_count = seq_count * self.SEQ_LEN
                    total_samples = seq_count
                    ws.append([idx, label, seq_count, frame_count, total_samples])
            
            # Summary
            total_labels = len(self.labels_order)
            total_seqs = sum(len(v) for v in self.dataset.values())
            total_frames = total_seqs * self.SEQ_LEN
            
            ws.append([])
            ws.append(['TỔNG CỘNG', '', total_seqs, total_frames, total_seqs])
            
            # Style summary
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='F0F9FF', end_color='F0F9FF', fill_type='solid')
            
            # Auto width
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            wb.save(filename)
            messagebox.showinfo("Thành công", f"Đã xuất Excel:\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xuất Excel: {e}")
    
    # ===== RECOGNITION =====
    
    def recognize_current(self):
        if not self.model:
            return
        
        seq = np.array(list(self.frame_buffer)).reshape(1, self.SEQ_LEN, 784)
        probs = self.model.predict(seq, verbose=0)[0]
        
        idx = int(np.argmax(probs))
        conf = float(probs[idx])
        
        labels = [lb for lb in self.labels_order if lb in self.dataset and len(self.dataset[lb]) > 0]
        label = labels[idx] if idx < len(labels) else "—"
        
        if conf > 0.6:
            self.lbl_pred.config(text=label)
            self.lbl_conf.config(text=f"Độ tin cậy: {conf*100:.1f}%")
            
            if self.tts:
                threading.Thread(target=lambda: self._speak(label), daemon=True).start()
        else:
            self.lbl_pred.config(text="❓")
            self.lbl_conf.config(text="Không chắc chắn")
    
    def _speak(self, text):
        """Speak text using TTS"""
        with self.tts_lock:
            try:
                self.tts.say(text)
                self.tts.runAndWait()
            except:
                pass
    
    # ===== SPEECH RECOGNITION =====
    
    def start_listening(self):
        if not sr or not self.recognizer:
            messagebox.showerror("Thiếu thư viện", "Cài: pip install SpeechRecognition pyaudio")
            return
        
        try:
            self.mic = sr.Microphone()
        except Exception as e:
            messagebox.showerror("Lỗi Mic", str(e))
            return
        
        self.stt_is_listening = True
        self.btn_listen.config(state=tk.DISABLED)
        self.btn_stop_listen.config(state=tk.NORMAL)
        self.lbl_mic.config(text="🎙️ Đang nghe...", fg=self.COLOR_SUCCESS)
        self.lock_interface(True)
        
        threading.Thread(target=self._listen_loop, daemon=True).start()
    
    def stop_listening(self):
        self.stt_is_listening = False
        self.btn_listen.config(state=tk.NORMAL)
        self.btn_stop_listen.config(state=tk.DISABLED)
        self.lbl_mic.config(text="🎤 Microphone: Sẵn sàng", fg=self.COLOR_TEXT)
        self.lock_interface(False)
        
        # Close playback
        self.playback_running = False
        self.canvas.delete("all")
    
    def _listen_loop(self):
        with self.mic as source:
            self.recognizer.adjust_for_ambient_noise(source, duration=0.6)
        
        while self.stt_is_listening:
            try:
                with self.mic as source:
                    audio = self.recognizer.listen(source, timeout=3, phrase_time_limit=5)
                
                try:
                    text = self.recognizer.recognize_google(audio, language='vi-VN')
                except:
                    text = ""
                
                if text:
                    self.root.after(0, lambda t=text: self.on_speech_text(t))
            except:
                continue
    
    def on_speech_text(self, text):
        self.txt_speech.delete('1.0', tk.END)
        self.txt_speech.insert(tk.END, text)
        
        best, score = self.best_label(text)
        self.lbl_match.config(text=f"Khớp: {best or '—'} ({int(score*100)}%)")
        
        if best and score > 0.5:
            self.play_label(best)
    
    def best_label(self, text):
        labels = [lb for lb in self.labels_order if lb in self.dataset and len(self.dataset[lb]) > 0]
        if not labels:
            return None, 0.0
        
        t = text.strip().lower()
        lst = [s.lower() for s in labels]
        
        m = difflib.get_close_matches(t, lst, n=1, cutoff=0.4)
        if m:
            idx = lst.index(m[0])
            return labels[idx], 1.0
        
        scores = [(lab, difflib.SequenceMatcher(None, t, lab.lower()).ratio()) for lab in labels]
        best = max(scores, key=lambda x: x[1])
        return best[0], float(best[1])
    
    def play_label(self, label):
        seq_dirs = self.frames_index.get(label, [])
        if not seq_dirs:
            return
        
        seq_rel = seq_dirs[-1]
        seq_dir = self.frames_root / seq_rel
        frames = sorted(seq_dir.glob("*.jpg"))
        
        if not frames:
            return
        
        self.playback_frames = [Image.open(p) for p in frames]
        self.playback_idx = 0
        self.playback_label = label
        self.playback_running = True
        
        if self.is_running:
            self.stop_camera()
        
        self.render_playback()
    
    def render_playback(self):
        if not self.playback_running or not self.playback_frames:
            return
        
        canvas_w = self.canvas.winfo_width() or 800
        canvas_h = self.canvas.winfo_height() or 600
        
        base = Image.new("RGB", (canvas_w, canvas_h), (26, 26, 26))
        
        w, h = int(canvas_w * 0.7), int(canvas_h * 0.7)
        x, y = (canvas_w - w) // 2, (canvas_h - h) // 2
        
        frame_img = self.playback_frames[self.playback_idx % len(self.playback_frames)]
        frame_img = frame_img.resize((w, h), Image.LANCZOS)
        base.paste(frame_img, (x, y))
        
        draw = ImageDraw.Draw(base)
        draw.rectangle((x, y - 45, x + 320, y - 5), fill=(59, 130, 246))
        draw.text((x + 15, y - 35), f"📹 {self.playback_label}", fill=(255, 255, 255))
        
        photo = ImageTk.PhotoImage(base)
        self.canvas.create_image(0, 0, anchor=tk.NW, image=photo)
        self.canvas.image = photo
        
        self.playback_idx = (self.playback_idx + 1) % len(self.playback_frames)
        self.root.after(60, self.render_playback)
    
    # ===== UTILS =====
    
    def update_stats(self):
        tot_labels = len([lb for lb in self.dataset if len(self.dataset[lb]) > 0])
        tot_seq = sum(len(v) for v in self.dataset.values())
        
        if self.current_mode == "collect":
            label = self.entry_label.get().strip()
            cur = len(self.dataset.get(label, [])) if label else 0
            self.lbl_collect_stats.config(
                text=f"Sequences: {cur} | Tổng: {tot_seq}"
            )
        
        model_status = f"✅ Sẵn sàng" if self.model else "Chưa train"
        self.lbl_train_stats.config(
            text=f"Dataset: {tot_labels} nhãn, {tot_seq} sequences\nMô hình: {model_status}"
        )
    
    def save_dataset(self):
        with open(self.dataset_pkl, 'wb') as f:
            pickle.dump({
                'dataset': self.dataset,
                'labels_order': self.labels_order
            }, f)
    
    def load_dataset(self):
        if self.dataset_pkl.exists():
            try:
                obj = pickle.load(open(self.dataset_pkl, 'rb'))
                if isinstance(obj, dict) and 'dataset' in obj:
                    self.dataset = obj.get('dataset', {})
                    self.labels_order = obj.get('labels_order', list(self.dataset.keys()))
                else:
                    self.dataset = obj
                    self.labels_order = list(self.dataset.keys())
            except:
                self.dataset = {}
                self.labels_order = []
    
    def save_frames_index(self):
        with open(self.frames_index_path, 'w', encoding='utf-8') as f:
            json.dump(self.frames_index, f, ensure_ascii=False, indent=2)
    
    def load_frames_index(self):
        if self.frames_index_path.exists():
            try:
                self.frames_index = json.load(
                    open(self.frames_index_path, 'r', encoding='utf-8')
                )
            except:
                self.frames_index = {}


if __name__ == "__main__":
    print("\n" + "="*80)
    print("🤖 ROBOT PHIÊN DỊCH NGÔN NGỮ KÝ HIỆU 2 CHIỀU")
    print("Ứng dụng học máy Deep Learning - ULTIMATE VERSION")
    print("="*80)
    print("\n✨ TÍNH NĂNG HOÀN CHỈNH:")
    print("  ✅ Giao diện siêu đẹp với màu sắc hiện đại")
    print("  🔒 Bảo mật Thu thập & Train bằng password (admin123)")
    print("  📊 Xuất thống kê ra Excel")
    print("  📈 Hiển thị % training với progress bar")
    print("  🤖 Animation robot học tập")
    print("  🎯 Nhận diện cử chỉ → TTS")
    print("  🎙️ Giọng nói → Animation")
    print("\n" + "="*80 + "\n")
    
    root = tk.Tk()
    app = UltimateSignLanguageApp(root)
    root.mainloop()
