"""
🤖 ROBOT PHIÊN DỊCH NGÔN NGỮ KÝ HIỆU 2 CHIỀU v2.1
Ứng dụng học máy Deep Learning - OPTIMIZED VERSION

Cải tiến:
- Password session-based (chỉ nhập 1 lần)
- TTS tiếng Việt online (Google TTS)
- Fix TTS nói nhiều lần
- Fix giao diện crash
- Nút xóa rõ ràng hơn
- Tối ưu performance
- Thanh cuộn dọc
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog, simpledialog
import cv2
import numpy as np
import pickle, json, os, time
from datetime import datetime
from PIL import Image, ImageTk, ImageDraw, ImageFont
import threading
from collections import deque
import difflib
from pathlib import Path

os.environ.setdefault('TF_CPP_MIN_LOG_LEVEL','2')
import tensorflow as tf
from tensorflow import keras

try:
    import speech_recognition as sr
except:
    sr = None

# Google TTS
try:
    from gtts import gTTS
    import pygame
    GTTS_AVAILABLE = True
except:
    GTTS_AVAILABLE = False
    print("⚠️ Google TTS không khả dụng. Cài: pip install gTTS pygame")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except:
    openpyxl = None


class UltimateSignLanguageApp:
    SEQ_LEN = 30
    PASSWORD = "admin123"
    
    # Color scheme
    COLOR_PRIMARY = '#3b82f6'
    COLOR_SUCCESS = '#10b981'
    COLOR_WARNING = '#f59e0b'
    COLOR_DANGER = '#ef4444'
    COLOR_PURPLE = '#8b5cf6'
    COLOR_BG = '#f0f4f8'
    COLOR_CARD = '#ffffff'
    COLOR_TEXT = '#1e293b'
    COLOR_TEXT_LIGHT = '#64748b'
    COLOR_BORDER = '#cbd5e1'
    COLOR_ACCENT = '#06b6d4'
    
    def __init__(self, root):
        self.root = root
        self.root.title("🤖 Robot Phiên Dịch Ngôn Ngữ Ký Hiệu 2 Chiều v2.1 - Deep Learning")
        self.root.geometry("1450x920")
        self.root.configure(bg=self.COLOR_BG)
        
        # Session-based authentication
        self.authenticated = False
        
        # State
        self.current_mode = None
        self.is_locked = False
        
        # Camera
        self.cap = None
        self.is_running = False
        self.frame_buffer = deque(maxlen=self.SEQ_LEN)
        self.raw_roi_buffer = deque(maxlen=self.SEQ_LEN)
        
        # Collection
        self.is_collecting = False
        self.current_label = ""
        
        # Data
        self.dataset = {}
        self.labels_order = []
        self.model = None
        self.model_path = Path("sign_language_model.h5")
        
        self.dataset_pkl = Path("dataset_dynamic.pkl")
        self.frames_root = Path("frames")
        self.frames_root.mkdir(exist_ok=True)
        self.frames_index_path = Path("frames_index.json")
        self.frames_index = {}
        
        self.lower_skin = np.array([0, 20, 70], dtype=np.uint8)
        self.upper_skin = np.array([20, 255, 255], dtype=np.uint8)
        
        # Speech
        self.recognizer = sr.Recognizer() if sr else None
        self.stt_is_listening = False
        
        # TTS - Google TTS
        self.tts_queue = []
        self.tts_playing = False
        self.tts_lock = threading.Lock()
        self.last_spoken = None
        self.last_spoken_time = 0
        self.gtts_available = GTTS_AVAILABLE
        
        if self.gtts_available:
            try:
                pygame.mixer.init(frequency=22050, size=-16, channels=2, buffer=512)
                print("✅ Pygame mixer initialized for TTS")
            except Exception as e:
                print(f"⚠️ Pygame mixer init warning: {e}")
                self.gtts_available = False
        
        # Training progress
        self.training_progress = 0
        self.training_status = ""
        
        # Playback
        self.playback_frames = []
        self.playback_idx = 0
        self.playback_running = False
        self.playback_label = None
        
        # Robot animation
        self.robot_frame_idx = 0
        self.robot_animating = False
        
        # Performance optimization
        self.frame_skip = 0
        self.recognition_cooldown = 0
        
        # Load data
        self.load_dataset()
        self.load_frames_index()
        self.load_model()  # Load existing model if available
        
        # Build UI
        self.build_ui()
        
        print("\n✅ 🤖 ROBOT PHIÊN DỊCH v2.1 READY!\n")
    
    def build_ui(self):
        # ===== TOP HEADER =====
        header = tk.Frame(self.root, bg=self.COLOR_PRIMARY, height=100)
        header.pack(fill=tk.X, side=tk.TOP)
        header.pack_propagate(False)
        
        icon_frame = tk.Frame(header, bg=self.COLOR_PRIMARY)
        icon_frame.pack(side=tk.LEFT, padx=30, pady=20)
        
        tk.Label(
            icon_frame,
            text="🤖",
            font=('Arial', 48),
            bg=self.COLOR_PRIMARY,
            fg='white'
        ).pack()
        
        title_frame = tk.Frame(header, bg=self.COLOR_PRIMARY)
        title_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=15)
        
        tk.Label(
            title_frame,
            text="ROBOT PHIÊN DỊCH NGÔN NGỮ KÝ HIỆU 2 CHIỀU v2.1",
            font=('Arial', 22, 'bold'),
            bg=self.COLOR_PRIMARY,
            fg='white'
        ).pack(anchor=tk.W)
        
        tk.Label(
            title_frame,
            text="Ứng dụng học máy Deep Learning - LSTM + Google TTS",
            font=('Arial', 11),
            bg=self.COLOR_PRIMARY,
            fg='#bfdbfe'
        ).pack(anchor=tk.W, pady=(5, 0))
        
        version_frame = tk.Frame(header, bg=self.COLOR_PRIMARY)
        version_frame.pack(side=tk.RIGHT, padx=30)
        
        tk.Label(
            version_frame,
            text="v2.1",
            font=('Arial', 10, 'bold'),
            bg='#1e40af',
            fg='white',
            padx=12,
            pady=5
        ).pack()
        
        tk.Label(
            version_frame,
            text="🔒 Protected",
            font=('Arial', 9),
            bg=self.COLOR_PRIMARY,
            fg='#bfdbfe'
        ).pack(pady=(5, 0))
        
        # ===== MODE SELECTOR =====
        mode_container = tk.Frame(self.root, bg=self.COLOR_BG)
        mode_container.pack(fill=tk.X, pady=20, padx=25)
        
        tk.Label(
            mode_container,
            text="⚡ Chọn chế độ hoạt động:",
            font=('Arial', 13, 'bold'),
            bg=self.COLOR_BG,
            fg=self.COLOR_TEXT
        ).pack(side=tk.LEFT, padx=10)
        
        self.mode_buttons = {}
        modes = [
            ("collect", "✋ Thu thập", self.COLOR_SUCCESS, "🔒"),
            ("train", "🧠 Huấn luyện", self.COLOR_PURPLE, "🔒"),
            ("recognize", "🎯 Nhận diện", self.COLOR_PRIMARY, ""),
            ("speech", "🎙️ Giọng nói", self.COLOR_WARNING, "")
        ]
        
        for mode_id, text, color, lock in modes:
            btn_text = f"{text} {lock}" if lock else text
            btn = tk.Button(
                mode_container,
                text=btn_text,
                font=('Arial', 11, 'bold'),
                bg=color,
                fg='white',
                relief=tk.FLAT,
                padx=22,
                pady=12,
                cursor='hand2',
                command=lambda m=mode_id: self.show_mode(m)
            )
            btn.pack(side=tk.LEFT, padx=5)
            self.mode_buttons[mode_id] = btn
        
        # ===== MAIN CONTAINER =====
        main = tk.Frame(self.root, bg=self.COLOR_BG)
        main.pack(fill=tk.BOTH, expand=True, padx=25, pady=(0, 20))
        
        # Left: Display
        left = tk.Frame(main, bg=self.COLOR_CARD, relief=tk.SOLID, bd=1)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 12))
        
        disp_header = tk.Frame(left, bg=self.COLOR_CARD, height=50)
        disp_header.pack(fill=tk.X)
        disp_header.pack_propagate(False)
        
        tk.Label(
            disp_header,
            text="📺 Màn hình hiển thị",
            font=('Arial', 13, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT
        ).pack(side=tk.LEFT, padx=20, pady=15)
        
        self.lbl_mode_indicator = tk.Label(
            disp_header,
            text="",
            font=('Arial', 11, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_PRIMARY
        )
        self.lbl_mode_indicator.pack(side=tk.RIGHT, padx=20)
        
        tk.Frame(left, bg=self.COLOR_BORDER, height=1).pack(fill=tk.X)
        
        # Canvas
        canvas_container = tk.Frame(left, bg='black', relief=tk.SUNKEN, bd=2)
        canvas_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
        
        self.canvas = tk.Canvas(canvas_container, bg='#1a1a1a', highlightthickness=0)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        
        # Camera controls
        self.cam_controls = tk.Frame(left, bg=self.COLOR_CARD)
        self.cam_controls.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        cam_btn_frame = tk.Frame(self.cam_controls, bg=self.COLOR_CARD)
        cam_btn_frame.pack(side=tk.LEFT)
        
        self.btn_cam_on = tk.Button(
            cam_btn_frame,
            text="📷 Bật Camera",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_PRIMARY,
            fg='white',
            relief=tk.FLAT,
            padx=18,
            pady=9,
            cursor='hand2',
            command=self.start_camera
        )
        self.btn_cam_on.pack(side=tk.LEFT, padx=3)
        
        self.btn_cam_off = tk.Button(
            cam_btn_frame,
            text="⏹ Tắt",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_DANGER,
            fg='white',
            relief=tk.FLAT,
            padx=18,
            pady=9,
            cursor='hand2',
            command=self.stop_camera,
            state=tk.DISABLED
        )
        self.btn_cam_off.pack(side=tk.LEFT, padx=3)
        
        cam_status = tk.Frame(self.cam_controls, bg=self.COLOR_CARD)
        cam_status.pack(side=tk.RIGHT)
        
        self.lbl_cam_status = tk.Label(
            cam_status,
            text="⚪ Tắt",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT_LIGHT
        )
        self.lbl_cam_status.pack(side=tk.LEFT, padx=10)
        
        self.lbl_buffer = tk.Label(
            cam_status,
            text="Buffer: 0/30",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_WARNING
        )
        self.lbl_buffer.pack(side=tk.LEFT, padx=10)
        
        # Right: Control Panel with SCROLLBAR
        right = tk.Frame(main, bg=self.COLOR_CARD, relief=tk.SOLID, bd=1, width=520)
        right.pack(side=tk.RIGHT, fill=tk.BOTH, padx=(12, 0))
        right.pack_propagate(False)
        
        # Control header (fixed at top)
        ctrl_header = tk.Frame(right, bg=self.COLOR_CARD, height=50)
        ctrl_header.pack(fill=tk.X)
        ctrl_header.pack_propagate(False)
        
        self.lbl_panel_title = tk.Label(
            ctrl_header,
            text="⚙️ Bảng điều khiển",
            font=('Arial', 13, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT
        )
        self.lbl_panel_title.pack(pady=15)
        
        # Separator
        tk.Frame(right, bg=self.COLOR_BORDER, height=1).pack(fill=tk.X)
        
        # Scrollable container
        scroll_container = tk.Frame(right, bg=self.COLOR_CARD)
        scroll_container.pack(fill=tk.BOTH, expand=True)
        
        # Create canvas with scrollbar
        self.control_canvas = tk.Canvas(scroll_container, bg=self.COLOR_CARD, highlightthickness=0)
        scrollbar = ttk.Scrollbar(scroll_container, orient=tk.VERTICAL, command=self.control_canvas.yview)
        
        # Scrollable frame inside canvas
        self.panels_container = tk.Frame(self.control_canvas, bg=self.COLOR_CARD)
        
        # Configure canvas
        self.control_canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack scrollbar and canvas
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.control_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Create window in canvas
        self.canvas_window = self.control_canvas.create_window((0, 0), window=self.panels_container, anchor=tk.NW)
        
        # Bind configure event to update scroll region
        self.panels_container.bind('<Configure>', self._on_panel_configure)
        self.control_canvas.bind('<Configure>', self._on_canvas_configure)
        
        # Bind mousewheel for scrolling
        self.control_canvas.bind_all('<MouseWheel>', self._on_mousewheel)
        self.control_canvas.bind_all('<Button-4>', self._on_mousewheel)  # Linux scroll up
        self.control_canvas.bind_all('<Button-5>', self._on_mousewheel)  # Linux scroll down
        
        # Build all panels
        self.build_collect_panel()
        self.build_train_panel()
        self.build_recognize_panel()
        self.build_speech_panel()
        
        # Hide all initially
        self.collect_panel.pack_forget()
        self.train_panel.pack_forget()
        self.recognize_panel.pack_forget()
        self.speech_panel.pack_forget()
        
        self.show_welcome()
    
    def _on_panel_configure(self, event=None):
        """Update scroll region when panels change"""
        self.control_canvas.configure(scrollregion=self.control_canvas.bbox("all"))
    
    def _on_canvas_configure(self, event):
        """Update canvas window width when canvas is resized"""
        canvas_width = event.width
        self.control_canvas.itemconfig(self.canvas_window, width=canvas_width)
    
    def _on_mousewheel(self, event):
        """Handle mousewheel scrolling"""
        try:
            if event.num == 5 or event.delta < 0:
                self.control_canvas.yview_scroll(1, "units")
            elif event.num == 4 or event.delta > 0:
                self.control_canvas.yview_scroll(-1, "units")
        except:
            pass
    
    def show_welcome(self):
        """Show welcome screen"""
        welcome = tk.Frame(self.panels_container, bg=self.COLOR_CARD)
        welcome.pack(fill=tk.BOTH, expand=True, pady=50)
        
        tk.Label(
            welcome,
            text="🤖",
            font=('Arial', 80),
            bg=self.COLOR_CARD
        ).pack(pady=20)
        
        tk.Label(
            welcome,
            text="Chào mừng đến với\nRobot Phiên Dịch v2.1!",
            font=('Arial', 16, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT,
            justify=tk.CENTER
        ).pack(pady=10)
        
        tk.Label(
            welcome,
            text="Chọn chế độ bên trên để bắt đầu",
            font=('Arial', 11),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT_LIGHT
        ).pack()
        
        # Update scroll region
        self.root.after(100, self._on_panel_configure)
    
    def build_collect_panel(self):
        self.collect_panel = tk.Frame(self.panels_container, bg=self.COLOR_CARD)
        
        # Input section
        input_section = tk.Frame(self.collect_panel, bg='#f8fafc', relief=tk.SOLID, bd=1)
        input_section.pack(fill=tk.X, padx=20, pady=15)
        
        tk.Label(
            input_section,
            text="📝 Tên động tác:",
            font=('Arial', 10, 'bold'),
            bg='#f8fafc',
            fg=self.COLOR_TEXT
        ).pack(pady=(10, 5), padx=15, anchor=tk.W)
        
        self.entry_label = tk.Entry(
            input_section,
            font=('Arial', 11),
            bg='white',
            fg=self.COLOR_TEXT,
            relief=tk.SOLID,
            bd=1,
            insertbackground=self.COLOR_TEXT
        )
        self.entry_label.pack(fill=tk.X, padx=15, pady=(0, 10), ipady=7)
        
        # Buttons
        btn_frame = tk.Frame(self.collect_panel, bg=self.COLOR_CARD)
        btn_frame.pack(pady=12)
        
        self.btn_start_collect = tk.Button(
            btn_frame,
            text="🎬 Bắt đầu ghi",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_SUCCESS,
            fg='white',
            relief=tk.FLAT,
            padx=18,
            pady=10,
            cursor='hand2',
            command=self.start_collecting
        )
        self.btn_start_collect.pack(side=tk.LEFT, padx=4)
        
        self.btn_stop_collect = tk.Button(
            btn_frame,
            text="⏸ Dừng",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_WARNING,
            fg='white',
            relief=tk.FLAT,
            padx=18,
            pady=10,
            cursor='hand2',
            command=self.stop_collecting,
            state=tk.DISABLED
        )
        self.btn_stop_collect.pack(side=tk.LEFT, padx=4)
        
        # Stats
        stats_frame = tk.Frame(self.collect_panel, bg='#f0f9ff', relief=tk.SOLID, bd=1)
        stats_frame.pack(fill=tk.X, padx=20, pady=(0, 15))
        
        self.lbl_collect_stats = tk.Label(
            stats_frame,
            text="Sequences: 0 | Tổng: 0",
            font=('Arial', 10, 'bold'),
            bg='#f0f9ff',
            fg=self.COLOR_PRIMARY
        )
        self.lbl_collect_stats.pack(pady=10)
        
        # Sequences list
        tk.Label(
            self.collect_panel,
            text="📋 Danh sách sequences:",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT
        ).pack(pady=(5, 5), padx=20, anchor=tk.W)
        
        list_frame = tk.Frame(self.collect_panel, bg=self.COLOR_CARD)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 10))
        
        list_scrollbar = tk.Scrollbar(list_frame)
        list_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.seq_listbox = tk.Listbox(
            list_frame,
            font=('Consolas', 9),
            bg='#f8fafc',
            fg=self.COLOR_TEXT,
            relief=tk.SOLID,
            bd=1,
            selectmode=tk.SINGLE,
            yscrollcommand=list_scrollbar.set,
            height=8
        )
        self.seq_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        list_scrollbar.config(command=self.seq_listbox.yview)
        
        # Delete buttons - HIGHLY VISIBLE
        del_warning = tk.Frame(self.collect_panel, bg=self.COLOR_CARD)
        del_warning.pack(fill=tk.X, padx=20, pady=(10, 5))
        
        tk.Label(
            del_warning,
            text="⚠️ QUẢN LÝ XÓA DỮ LIỆU:",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_DANGER
        ).pack(anchor=tk.W)
        
        del_btn_frame = tk.Frame(self.collect_panel, bg='#fee2e2', relief=tk.RIDGE, bd=3)
        del_btn_frame.pack(fill=tk.X, padx=20, pady=(0, 15))
        
        btn_container = tk.Frame(del_btn_frame, bg='#fee2e2')
        btn_container.pack(pady=10, padx=10)
        
        tk.Button(
            btn_container,
            text="🗑️  XÓA SEQUENCE ĐÃ CHỌN",
            font=('Arial', 9, 'bold'),
            bg=self.COLOR_DANGER,
            fg='white',
            activebackground='#dc2626',
            activeforeground='white',
            relief=tk.RAISED,
            bd=3,
            padx=15,
            pady=8,
            cursor='hand2',
            command=self.delete_selected_sequence
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            btn_container,
            text="🗑️  XÓA TOÀN BỘ LABEL",
            font=('Arial', 9, 'bold'),
            bg='#991b1b',
            fg='white',
            activebackground='#7f1d1d',
            activeforeground='white',
            relief=tk.RAISED,
            bd=3,
            padx=15,
            pady=8,
            cursor='hand2',
            command=self.delete_entire_label
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Label(
            del_btn_frame,
            text="💡 Chọn sequence trong danh sách trước khi xóa",
            font=('Arial', 8, 'italic'),
            bg='#fee2e2',
            fg='#7f1d1d'
        ).pack(pady=(0, 8))
    
    def build_train_panel(self):
        self.train_panel = tk.Frame(self.panels_container, bg=self.COLOR_CARD)
        
        # Stats card
        stats_card = tk.Frame(self.train_panel, bg='#f0f9ff', relief=tk.SOLID, bd=1)
        stats_card.pack(fill=tk.X, padx=20, pady=15)
        
        self.lbl_train_stats = tk.Label(
            stats_card,
            text="Dataset: 0 nhãn, 0 sequences\nMô hình: Chưa train",
            font=('Arial', 11, 'bold'),
            bg='#f0f9ff',
            fg=self.COLOR_TEXT,
            justify=tk.CENTER
        )
        self.lbl_train_stats.pack(pady=15)
        
        # Progress bar
        self.progress_frame = tk.Frame(self.train_panel, bg=self.COLOR_CARD)
        
        tk.Label(
            self.progress_frame,
            text="📊 Tiến độ huấn luyện:",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT
        ).pack(pady=(0, 5))
        
        style = ttk.Style()
        style.configure("Custom.Horizontal.TProgressbar",
                       troughcolor='#e2e8f0',
                       background=self.COLOR_SUCCESS,
                       thickness=20)
        
        self.train_progress = ttk.Progressbar(
            self.progress_frame,
            length=400,
            mode='determinate',
            style='Custom.Horizontal.TProgressbar'
        )
        self.train_progress.pack(pady=5)
        
        self.lbl_train_progress = tk.Label(
            self.progress_frame,
            text="0%",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_PRIMARY
        )
        self.lbl_train_progress.pack(pady=5)
        
        # Robot animation
        self.robot_canvas = tk.Canvas(
            self.progress_frame,
            width=100,
            height=100,
            bg=self.COLOR_CARD,
            highlightthickness=0
        )
        self.robot_canvas.pack(pady=10)
        
        # Buttons
        btn_frame = tk.Frame(self.train_panel, bg=self.COLOR_CARD)
        btn_frame.pack(pady=15)
        
        self.btn_train = tk.Button(
            btn_frame,
            text="🚀 Huấn luyện mô hình",
            font=('Arial', 11, 'bold'),
            bg=self.COLOR_PURPLE,
            fg='white',
            relief=tk.FLAT,
            padx=25,
            pady=13,
            cursor='hand2',
            command=self.train_model
        )
        self.btn_train.pack(pady=6)
        
        tk.Button(
            btn_frame,
            text="📊 Xuất Excel",
            font=('Arial', 11, 'bold'),
            bg=self.COLOR_ACCENT,
            fg='white',
            relief=tk.FLAT,
            padx=25,
            pady=13,
            cursor='hand2',
            command=self.export_to_excel
        ).pack(pady=6)
        
        tk.Button(
            btn_frame,
            text="🗑️ Xóa dataset",
            font=('Arial', 11, 'bold'),
            bg=self.COLOR_DANGER,
            fg='white',
            relief=tk.FLAT,
            padx=25,
            pady=13,
            cursor='hand2',
            command=self.clear_dataset
        ).pack(pady=6)
        
        # Statistics table
        tk.Label(
            self.train_panel,
            text="📊 Thống kê chi tiết:",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT
        ).pack(pady=(5, 8), padx=20, anchor=tk.W)
        
        tree_frame = tk.Frame(self.train_panel, bg=self.COLOR_CARD)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 15))
        
        tree_scroll = tk.Scrollbar(tree_frame)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        style.configure("Custom.Treeview", 
                       background='#f8fafc',
                       foreground=self.COLOR_TEXT,
                       rowheight=25,
                       fieldbackground='#f8fafc')
        style.configure("Custom.Treeview.Heading",
                       background=self.COLOR_PRIMARY,
                       foreground='white',
                       font=('Arial', 9, 'bold'))
        
        self.stats_tree = ttk.Treeview(
            tree_frame,
            columns=('label', 'sequences', 'frames'),
            show='headings',
            height=6,
            yscrollcommand=tree_scroll.set,
            style="Custom.Treeview"
        )
        tree_scroll.config(command=self.stats_tree.yview)
        
        self.stats_tree.heading('label', text='Label')
        self.stats_tree.heading('sequences', text='Sequences')
        self.stats_tree.heading('frames', text='Frames')
        
        self.stats_tree.column('label', width=180, anchor=tk.W)
        self.stats_tree.column('sequences', width=90, anchor=tk.CENTER)
        self.stats_tree.column('frames', width=90, anchor=tk.CENTER)
        
        self.stats_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    
    def build_recognize_panel(self):
        self.recognize_panel = tk.Frame(self.panels_container, bg=self.COLOR_CARD)
        
        tk.Label(
            self.recognize_panel,
            text="🎯 Kết quả nhận diện:",
            font=('Arial', 11, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT
        ).pack(pady=15)
        
        # Result display
        result_frame = tk.Frame(
            self.recognize_panel,
            bg='#eff6ff',
            relief=tk.SOLID,
            bd=2
        )
        result_frame.pack(fill=tk.X, padx=25, pady=10)
        
        self.lbl_pred = tk.Label(
            result_frame,
            text="—",
            font=('Arial', 52, 'bold'),
            bg='#eff6ff',
            fg=self.COLOR_PRIMARY
        )
        self.lbl_pred.pack(pady=25)
        
        self.lbl_conf = tk.Label(
            result_frame,
            text="Chờ buffer đầy...",
            font=('Arial', 12),
            bg='#eff6ff',
            fg=self.COLOR_TEXT_LIGHT
        )
        self.lbl_conf.pack(pady=(0, 25))
        
        # Info
        info_frame = tk.Frame(self.recognize_panel, bg='#f0fdf4', relief=tk.SOLID, bd=1)
        info_frame.pack(fill=tk.X, padx=25, pady=20)
        
        tts_title = "🔊 Google Text-to-Speech (Tiếng Việt)" if GTTS_AVAILABLE else "🔊 Text-to-Speech (Không khả dụng)"
        tts_color = self.COLOR_SUCCESS if GTTS_AVAILABLE else self.COLOR_DANGER
        
        tk.Label(
            info_frame,
            text=tts_title,
            font=('Arial', 10, 'bold'),
            bg='#f0fdf4',
            fg=tts_color
        ).pack(anchor=tk.W, padx=15, pady=(12, 5))
        
        tts_desc = "Robot sẽ tự động đọc tên động tác\nmỗi khi nhận diện thành công" if GTTS_AVAILABLE else "Cài đặt: pip install gTTS pygame\nđể kích hoạt tính năng này"
        
        tk.Label(
            info_frame,
            text=tts_desc,
            font=('Arial', 9),
            bg='#f0fdf4',
            fg=self.COLOR_TEXT_LIGHT,
            justify=tk.LEFT
        ).pack(anchor=tk.W, padx=15, pady=(0, 12))
    
    def build_speech_panel(self):
        self.speech_panel = tk.Frame(self.panels_container, bg=self.COLOR_CARD)
        
        # Mic status
        mic_frame = tk.Frame(self.speech_panel, bg='#fef3c7', relief=tk.SOLID, bd=1)
        mic_frame.pack(fill=tk.X, padx=20, pady=15)
        
        self.lbl_mic = tk.Label(
            mic_frame,
            text="🎤 Microphone: Sẵn sàng",
            font=('Arial', 11, 'bold'),
            bg='#fef3c7',
            fg=self.COLOR_TEXT
        )
        self.lbl_mic.pack(pady=12)
        
        # Buttons
        btn_frame = tk.Frame(self.speech_panel, bg=self.COLOR_CARD)
        btn_frame.pack(pady=12)
        
        self.btn_listen = tk.Button(
            btn_frame,
            text="🎧 Bắt đầu nghe",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_SUCCESS,
            fg='white',
            relief=tk.FLAT,
            padx=18,
            pady=10,
            cursor='hand2',
            command=self.start_listening
        )
        self.btn_listen.pack(side=tk.LEFT, padx=4)
        
        self.btn_stop_listen = tk.Button(
            btn_frame,
            text="🛑 Dừng",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_DANGER,
            fg='white',
            relief=tk.FLAT,
            padx=18,
            pady=10,
            cursor='hand2',
            command=self.stop_listening,
            state=tk.DISABLED
        )
        self.btn_stop_listen.pack(side=tk.LEFT, padx=4)
        
        # Text display
        tk.Label(
            self.speech_panel,
            text="📝 Văn bản nhận dạng:",
            font=('Arial', 10, 'bold'),
            bg=self.COLOR_CARD,
            fg=self.COLOR_TEXT
        ).pack(pady=(15, 5), padx=20, anchor=tk.W)
        
        text_frame = tk.Frame(self.speech_panel, bg=self.COLOR_CARD)
        text_frame.pack(fill=tk.X, padx=20, pady=5)
        
        self.txt_speech = tk.Text(
            text_frame,
            height=3,
            font=('Arial', 11),
            bg='#f8fafc',
            fg=self.COLOR_TEXT,
            relief=tk.SOLID,
            bd=1,
            wrap=tk.WORD
        )
        self.txt_speech.pack(fill=tk.X, ipady=4)
        
        # Match result
        match_frame = tk.Frame(self.speech_panel, bg='#f0f9ff', relief=tk.SOLID, bd=1)
        match_frame.pack(fill=tk.X, padx=20, pady=15)
        
        self.lbl_match = tk.Label(
            match_frame,
            text="Khớp: —",
            font=('Arial', 13, 'bold'),
            bg='#f0f9ff',
            fg=self.COLOR_PRIMARY
        )
        self.lbl_match.pack(pady=12)
        
        # Info
        info_frame = tk.Frame(self.speech_panel, bg='#fef2f2', relief=tk.SOLID, bd=1)
        info_frame.pack(fill=tk.X, padx=20, pady=(0, 15))
        
        tk.Label(
            info_frame,
            text="🔹 Animation Playback",
            font=('Arial', 10, 'bold'),
            bg='#fef2f2',
            fg=self.COLOR_DANGER
        ).pack(anchor=tk.W, padx=15, pady=(12, 5))
        
        tk.Label(
            info_frame,
            text="Khi tìm thấy label khớp,\nrobot sẽ phát animation tự động",
            font=('Arial', 9),
            bg='#fef2f2',
            fg=self.COLOR_TEXT_LIGHT,
            justify=tk.LEFT
        ).pack(anchor=tk.W, padx=15, pady=(0, 12))
    
    # ===== SESSION-BASED PASSWORD =====
    
    def check_password(self, action_name):
        """Session-based password check"""
        if self.authenticated:
            return True
        
        password = simpledialog.askstring(
            "🔒 Xác thực phiên làm việc",
            f"Nhập mật khẩu để truy cập {action_name}:\n(Chỉ cần nhập 1 lần cho phiên này)",
            show='*'
        )
        
        if password == self.PASSWORD:
            self.authenticated = True
            messagebox.showinfo("✅ Thành công", "Xác thực thành công!\nBạn có thể truy cập tất cả chức năng.")
            return True
        elif password is not None:
            messagebox.showerror("❌ Lỗi", "Mật khẩu sai!")
        return False
    
    # ===== MODE MANAGEMENT =====
    
    def show_mode(self, mode):
        # Check password for protected modes (session-based)
        if mode in ["collect", "train"] and not self.check_password("chức năng bảo mật"):
            return
        
        if self.is_locked:
            messagebox.showwarning("Đang hoạt động", "Vui lòng dừng hoạt động hiện tại!")
            return
        
        self.current_mode = mode
        
        # Hide all panels
        self.collect_panel.pack_forget()
        self.train_panel.pack_forget()
        self.recognize_panel.pack_forget()
        self.speech_panel.pack_forget()
        
        # Stop playback
        self.playback_running = False
        
        # Reset scroll to top
        self.control_canvas.yview_moveto(0)
        
        # Update UI
        mode_info = {
            "collect": ("✋ THU THẬP MẪU", "📋 Quản lý thu thập", self.COLOR_SUCCESS),
            "train": ("🧠 HUẤN LUYỆN", "🚀 Huấn luyện & Thống kê", self.COLOR_PURPLE),
            "recognize": ("🎯 NHẬN DIỆN", "🎯 Nhận diện động tác", self.COLOR_PRIMARY),
            "speech": ("🎙️ GIỌNG NÓI", "🎙️ Chuyển giọng nói", self.COLOR_WARNING)
        }
        
        indicator, title, color = mode_info[mode]
        self.lbl_mode_indicator.config(text=indicator, fg=color)
        self.lbl_panel_title.config(text=title)
        
        # Show panel
        if mode == "collect":
            self.collect_panel.pack(fill=tk.BOTH, expand=True)
            self.refresh_sequences_list()
            self.update_stats()
        elif mode == "train":
            self.train_panel.pack(fill=tk.BOTH, expand=True)
            self.progress_frame.pack_forget()
            self.refresh_stats_table()
            self.update_stats()
        elif mode == "recognize":
            self.recognize_panel.pack(fill=tk.BOTH, expand=True)
        elif mode == "speech":
            self.speech_panel.pack(fill=tk.BOTH, expand=True)
        
        # Update scroll region after packing
        self.root.after(100, self._on_panel_configure)
    
    def lock_interface(self, lock=True):
        """Lock/unlock interface"""
        self.is_locked = lock
        state = tk.DISABLED if lock else tk.NORMAL
        
        for btn in self.mode_buttons.values():
            btn.config(state=state)
    
    # ===== CAMERA - OPTIMIZED =====
    
    def start_camera(self):
        if self.is_running:
            return
        
        self.cap = cv2.VideoCapture(0)
        if not self.cap.isOpened():
            messagebox.showerror("Lỗi", "Không mở được camera!")
            return
        
        self.is_running = True
        self.btn_cam_on.config(state=tk.DISABLED)
        self.btn_cam_off.config(state=tk.NORMAL)
        self.lbl_cam_status.config(text="🟢 Đang chạy", fg=self.COLOR_SUCCESS)
        self.frame_buffer.clear()
        
        self.playback_running = False
        self.frame_skip = 0
        self.update_frame()
    
    def stop_camera(self):
        if not self.is_running:
            return
        
        self.is_running = False
        if self.cap:
            self.cap.release()
        
        self.btn_cam_on.config(state=tk.NORMAL)
        self.btn_cam_off.config(state=tk.DISABLED)
        self.lbl_cam_status.config(text="⚪ Tắt", fg=self.COLOR_TEXT_LIGHT)
        self.canvas.delete("all")
        self.frame_buffer.clear()
    
    def update_frame(self):
        if not self.is_running:
            return
        
        ret, frame = self.cap.read()
        if not ret:
            self.root.after(10, self.update_frame)
            return
        
        frame = cv2.flip(frame, 1)
        cv2.rectangle(frame, (100, 100), (540, 380), (0, 255, 0), 2)
        cv2.putText(frame, "Dong tac o day", (110, 90),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
        
        roi = frame[100:380, 100:540]
        
        # IMPORTANT: Always save raw ROI BEFORE processing
        roi_copy = roi.copy()
        
        hsv = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
        mask = cv2.inRange(hsv, self.lower_skin, self.upper_skin)
        
        kernel = np.ones((3,3), np.uint8)
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel, iterations=2)
        mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel, iterations=2)
        mask = cv2.GaussianBlur(mask, (5, 5), 100)
        
        features = self.image_to_features(mask)
        if features is not None:
            self.frame_buffer.append(features)
            self.raw_roi_buffer.append(roi_copy)  # Use the copy saved earlier
        
        buffer_size = len(self.frame_buffer)
        self.lbl_buffer.config(text=f"Buffer: {buffer_size}/30")
        
        if buffer_size == self.SEQ_LEN:
            if self.is_collecting and self.current_label:
                self.collect_sequence()
            elif self.current_mode == "recognize" and self.model:
                if self.recognition_cooldown == 0:
                    self.recognize_current()
                    self.recognition_cooldown = 10
        
        if self.recognition_cooldown > 0:
            self.recognition_cooldown -= 1
        
        mask_3ch = cv2.cvtColor(mask, cv2.COLOR_GRAY2BGR)
        frame[100:380, 100:540] = cv2.addWeighted(roi, 0.7, mask_3ch, 0.3, 0)
        
        # Skip frames for performance
        self.frame_skip += 1
        if self.frame_skip % 2 == 0:
            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(frame_rgb)
            
            canvas_w = self.canvas.winfo_width() or 800
            canvas_h = self.canvas.winfo_height() or 600
            img = img.resize((canvas_w, canvas_h), Image.LANCZOS)
            
            photo = ImageTk.PhotoImage(img)
            self.canvas.create_image(0, 0, anchor=tk.NW, image=photo)
            self.canvas.image = photo
        
        self.root.after(10, self.update_frame)
    
    def image_to_features(self, mask):
        try:
            resized = cv2.resize(mask, (28, 28))
            return resized.flatten() / 255.0
        except:
            return None
    
    # ===== COLLECTION =====
    
    def start_collecting(self):
        label = self.entry_label.get().strip()
        if not label:
            messagebox.showwarning("Cảnh báo", "Vui lòng nhập tên động tác!")
            return
        
        if not self.is_running:
            messagebox.showwarning("Cảnh báo", "Vui lòng bật camera!")
            return
        
        self.current_label = label
        self.is_collecting = True
        
        if label not in self.dataset:
            self.dataset[label] = []
            if label not in self.labels_order:
                self.labels_order.append(label)
        
        self.frame_buffer.clear()
        self.raw_roi_buffer.clear()
        
        self.btn_start_collect.config(state=tk.DISABLED)
        self.btn_stop_collect.config(state=tk.NORMAL)
        self.lock_interface(True)
    
    def stop_collecting(self):
        self.is_collecting = False
        self.current_label = ""
        
        self.btn_start_collect.config(state=tk.NORMAL)
        self.btn_stop_collect.config(state=tk.DISABLED)
        self.lock_interface(False)
        
        # Save dataset and frames index
        self.save_dataset()
        self.save_frames_index()
        
        print(f"\n📊 Collection stopped. Current frames_index:")
        for label, seqs in self.frames_index.items():
            print(f"  '{label}': {len(seqs)} sequences")
            for seq_path in seqs:
                full_path = self.frames_root / seq_path
                if full_path.exists():
                    frame_count = len(list(full_path.glob("*.jpg")))
                    print(f"    - {seq_path}: {frame_count} frames")
                else:
                    print(f"    - {seq_path}: ❌ NOT FOUND")
        
        self.refresh_sequences_list()
        self.update_stats()
    
    def collect_sequence(self):
        sequence = np.array(list(self.frame_buffer))
        self.dataset[self.current_label].append(sequence)
        
        count = len(self.dataset[self.current_label])
        
        # Save frames BEFORE clearing buffer
        if len(self.raw_roi_buffer) == self.SEQ_LEN:
            self.save_frames_for_sequence(self.current_label, count - 1)
            print(f"✅ Saved sequence {count} for '{self.current_label}' with {len(self.raw_roi_buffer)} frames")
        else:
            print(f"⚠️ Warning: raw_roi_buffer only has {len(self.raw_roi_buffer)} frames, expected {self.SEQ_LEN}")
        
        self.update_stats()
        self.frame_buffer.clear()
        self.raw_roi_buffer.clear()
    
    def save_frames_for_sequence(self, label, seq_idx):
        """Save raw frames to disk - Unicode-safe version"""
        try:
            if len(self.raw_roi_buffer) != self.SEQ_LEN:
                print(f"❌ Cannot save frames: buffer has {len(self.raw_roi_buffer)} frames, need {self.SEQ_LEN}")
                return
            
            label_dir = self.frames_root / label
            label_dir.mkdir(exist_ok=True, parents=True)
            
            seq_dir = label_dir / f"seq_{seq_idx:04d}"
            seq_dir.mkdir(exist_ok=True, parents=True)
            
            print(f"\n💾 Saving frames to: {seq_dir.absolute()}")
            
            saved_count = 0
            for i, roi in enumerate(self.raw_roi_buffer):
                frame_path = seq_dir / f"frame_{i:03d}.jpg"
                
                # Debug: Check ROI data
                if roi is None or not isinstance(roi, np.ndarray) or roi.size == 0:
                    print(f"⚠️ Frame {i}: Invalid ROI data")
                    continue
                
                # Use imencode/imdecode for Unicode path support
                try:
                    # Encode image to memory buffer
                    success, encoded = cv2.imencode('.jpg', roi, [cv2.IMWRITE_JPEG_QUALITY, 95])
                    
                    if success:
                        # Write buffer to file using Python (supports Unicode)
                        with open(str(frame_path), 'wb') as f:
                            f.write(encoded.tobytes())
                        
                        # Verify file was created
                        if frame_path.exists() and frame_path.stat().st_size > 0:
                            saved_count += 1
                        else:
                            print(f"⚠️ Frame {i}: File not created or empty")
                    else:
                        print(f"⚠️ Frame {i}: cv2.imencode failed")
                        
                except Exception as e:
                    print(f"❌ Exception saving frame {i}: {e}")
            
            print(f"✅ Successfully saved {saved_count}/{self.SEQ_LEN} frames to {seq_dir.name}")
            
            if saved_count == 0:
                print(f"❌ NO FRAMES SAVED! This is a critical error.")
                return
            
            if label not in self.frames_index:
                self.frames_index[label] = []
            
            rel_path = f"{label}/seq_{seq_idx:04d}"
            self.frames_index[label].append(rel_path)
            
        except Exception as e:
            print(f"❌ Error saving frames: {e}")
            import traceback
            traceback.print_exc()
    
    def refresh_sequences_list(self):
        self.seq_listbox.delete(0, tk.END)
        
        label = self.entry_label.get().strip()
        if not label or label not in self.dataset:
            return
        
        sequences = self.dataset[label]
        for i in range(len(sequences)):
            self.seq_listbox.insert(tk.END, f"  Sequence {i+1}")
    
    def delete_selected_sequence(self):
        label = self.entry_label.get().strip()
        if not label or label not in self.dataset:
            messagebox.showwarning("Cảnh báo", "Chọn label!")
            return
        
        selection = self.seq_listbox.curselection()
        if not selection:
            messagebox.showwarning("Cảnh báo", "Chọn sequence trong danh sách!")
            return
        
        idx = selection[0]
        
        if not messagebox.askyesno("⚠️ Xác nhận xóa", 
            f"Bạn có chắc muốn xóa Sequence {idx+1}?\nHành động này không thể hoàn tác!"):
            return
        
        del self.dataset[label][idx]
        
        if label in self.frames_index and idx < len(self.frames_index[label]):
            seq_rel = self.frames_index[label][idx]
            seq_dir = self.frames_root / seq_rel
            try:
                if seq_dir.exists():
                    for p in seq_dir.glob("*.jpg"):
                        p.unlink()
                    seq_dir.rmdir()
            except:
                pass
            
            del self.frames_index[label][idx]
        
        if len(self.dataset[label]) == 0:
            del self.dataset[label]
            self.labels_order.remove(label)
            if label in self.frames_index:
                del self.frames_index[label]
        
        self.save_dataset()
        self.save_frames_index()
        self.refresh_sequences_list()
        self.update_stats()
        messagebox.showinfo("✅ Thành công", "Đã xóa sequence!")
    
    def delete_entire_label(self):
        label = self.entry_label.get().strip()
        if not label or label not in self.dataset:
            messagebox.showwarning("Cảnh báo", "Label không tồn tại!")
            return
        
        seq_count = len(self.dataset[label])
        if not messagebox.askyesno("⚠️⚠️ XÁC NHẬN XÓA TOÀN BỘ", 
            f"BẠN SẮP XÓA TOÀN BỘ LABEL:\n'{label}'\n\n"
            f"Số sequences: {seq_count}\n"
            f"Hành động này KHÔNG THỂ HOÀN TÁC!\n\n"
            f"Bạn có chắc chắn muốn tiếp tục?"):
            return
        
        del self.dataset[label]
        self.labels_order.remove(label)
        
        if label in self.frames_index:
            label_dir = self.frames_root / label
            try:
                if label_dir.exists():
                    for p in label_dir.rglob("*"):
                        if p.is_file():
                            p.unlink()
                    for p in sorted(label_dir.glob("*"), reverse=True):
                        if p.is_dir():
                            try:
                                p.rmdir()
                            except:
                                pass
                    label_dir.rmdir()
            except:
                pass
            
            del self.frames_index[label]
        
        self.save_dataset()
        self.save_frames_index()
        self.refresh_sequences_list()
        self.update_stats()
        self.entry_label.delete(0, tk.END)
        messagebox.showinfo("✅ Hoàn tất", f"Đã xóa toàn bộ label '{label}'")
    
    # ===== TRAINING =====
    
    def train_model(self):
        labels = [lb for lb in self.labels_order if lb in self.dataset and len(self.dataset[lb]) > 0]
        if len(labels) < 2:
            messagebox.showwarning("Cảnh báo", "Cần ít nhất 2 nhãn!")
            return
        
        self.lock_interface(True)
        self.btn_train.config(state=tk.DISABLED, text="⏳ Đang train...")
        
        self.progress_frame.pack(fill=tk.X, padx=20, pady=15, before=self.btn_train.master)
        self.train_progress['value'] = 0
        self.lbl_train_progress.config(text="0%")
        
        self.robot_animating = True
        self.animate_robot()
        
        threading.Thread(target=self._train_thread, daemon=True).start()
    
    def animate_robot(self):
        """Animate robot while training"""
        if not self.robot_animating:
            self.robot_canvas.delete("all")
            return
        
        self.robot_canvas.delete("all")
        
        frame = self.robot_frame_idx % 4
        
        # Body
        self.robot_canvas.create_oval(25, 30, 75, 80, fill='#3b82f6', outline='#1e40af', width=2)
        
        # Eyes (blink)
        if frame == 3:
            self.robot_canvas.create_line(35, 45, 45, 45, fill='white', width=3)
            self.robot_canvas.create_line(55, 45, 65, 45, fill='white', width=3)
        else:
            self.robot_canvas.create_oval(35, 40, 45, 50, fill='white')
            self.robot_canvas.create_oval(55, 40, 65, 50, fill='white')
        
        # Antenna
        antenna_y = 20 - (frame % 2) * 3
        self.robot_canvas.create_line(50, 30, 50, antenna_y, fill='#1e40af', width=2)
        self.robot_canvas.create_oval(47, antenna_y-5, 53, antenna_y+1, fill='#ef4444')
        
        # Book
        book_x = 30 + (frame * 3)
        self.robot_canvas.create_rectangle(book_x, 55, book_x+15, 65, fill='#f59e0b', outline='#d97706')
        
        self.robot_frame_idx += 1
        self.root.after(200, self.animate_robot)
    
    def _train_thread(self):
        try:
            labels = [lb for lb in self.labels_order if lb in self.dataset and len(self.dataset[lb]) > 0]
            xs, ys = [], []
            
            for idx, label in enumerate(labels):
                for seq in self.dataset[label]:
                    xs.append(seq)
                    one_hot = np.zeros(len(labels))
                    one_hot[idx] = 1
                    ys.append(one_hot)
            
            xs = np.array(xs)
            ys = np.array(ys)
            
            self.model = keras.Sequential([
                keras.layers.LSTM(64, return_sequences=True, input_shape=(self.SEQ_LEN, 784)),
                keras.layers.Dropout(0.3),
                keras.layers.LSTM(32),
                keras.layers.Dropout(0.2),
                keras.layers.Dense(32, activation='relu'),
                keras.layers.Dense(len(labels), activation='softmax')
            ])
            
            self.model.compile(
                optimizer='adam',
                loss='categorical_crossentropy',
                metrics=['accuracy']
            )
            
            class ProgressCallback(keras.callbacks.Callback):
                def __init__(self, app):
                    self.app = app
                    
                def on_epoch_end(self, epoch, logs=None):
                    progress = int((epoch + 1) / 50 * 100)
                    self.app.root.after(0, lambda: self.app.update_training_progress(progress))
            
            history = self.model.fit(
                xs, ys,
                epochs=50,
                batch_size=16,
                validation_split=0.2,
                verbose=0,
                callbacks=[ProgressCallback(self)]
            )
            
            acc = history.history['val_accuracy'][-1]
            
            # Save model after training
            self.root.after(0, lambda: self.save_model())
            
            self.root.after(0, lambda: self._on_train_complete(len(labels), len(xs), acc))
            
        except Exception as e:
            print(f"✗ Error: {e}")
            self.root.after(0, lambda: messagebox.showerror("Lỗi", f"Training thất bại: {e}"))
            self.root.after(0, lambda: self._cleanup_training())
    
    def update_training_progress(self, progress):
        """Update progress bar - SAFE"""
        try:
            self.train_progress['value'] = progress
            self.lbl_train_progress.config(text=f"{progress}%")
        except:
            pass
    
    def _on_train_complete(self, n_labels, n_seqs, acc):
        """Callback when training completes"""
        self.robot_animating = False
        time.sleep(0.3)
        
        self.lock_interface(False)
        self.btn_train.config(state=tk.NORMAL, text="🚀 Huấn luyện mô hình")
        self.update_stats()
        self.refresh_stats_table()
        
        messagebox.showinfo("Thành công!", 
            f"🎉 Mô hình đã được huấn luyện & lưu!\n\n"
            f"📊 Labels: {n_labels}\n"
            f"📋 Sequences: {n_seqs}\n"
            f"✅ Accuracy: {acc*100:.1f}%\n"
            f"💾 Đã lưu: sign_language_model.h5")
    
    def _cleanup_training(self):
        """Cleanup after training error"""
        self.robot_animating = False
        self.lock_interface(False)
        self.btn_train.config(state=tk.NORMAL, text="🚀 Huấn luyện mô hình")
    
    def clear_dataset(self):
        if not messagebox.askyesno("⚠️⚠️ XÁC NHẬN", 
            "XÓA TOÀN BỘ DATASET?\n\nHành động này không thể hoàn tác!"):
            return
        
        self.dataset = {}
        self.labels_order = []
        self.model = None
        self.frames_index = {}
        
        # Delete model file
        try:
            if self.model_path.exists():
                self.model_path.unlink()
                print(f"🗑️ Deleted model file: {self.model_path}")
        except Exception as e:
            print(f"⚠️ Could not delete model: {e}")
        
        try:
            if self.frames_root.exists():
                for p in self.frames_root.rglob("*"):
                    if p.is_file():
                        p.unlink()
                for p in sorted(self.frames_root.glob("*"), reverse=True):
                    if p.is_dir():
                        try:
                            p.rmdir()
                        except:
                            pass
        except:
            pass
        
        self.save_dataset()
        self.save_frames_index()
        self.update_stats()
        self.refresh_stats_table()
        messagebox.showinfo("✅ Hoàn tất", "Đã xóa toàn bộ dataset & model")
    
    def refresh_stats_table(self):
        """Refresh statistics table"""
        for item in self.stats_tree.get_children():
            self.stats_tree.delete(item)
        
        for label in self.labels_order:
            if label in self.dataset:
                seq_count = len(self.dataset[label])
                frame_count = seq_count * self.SEQ_LEN
                self.stats_tree.insert('', tk.END, values=(label, seq_count, frame_count))
    
    def export_to_excel(self):
        """Export to Excel"""
        if not openpyxl:
            messagebox.showerror("Lỗi", "Vui lòng cài: pip install openpyxl")
            return
        
        if len(self.labels_order) == 0:
            messagebox.showwarning("Cảnh báo", "Không có dữ liệu!")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"sign_language_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Statistics"
            
            headers = ['STT', 'Label', 'Sequences', 'Frames', 'Total Samples']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for idx, label in enumerate(self.labels_order, 1):
                if label in self.dataset:
                    seq_count = len(self.dataset[label])
                    frame_count = seq_count * self.SEQ_LEN
                    total_samples = seq_count
                    ws.append([idx, label, seq_count, frame_count, total_samples])
            
            total_labels = len(self.labels_order)
            total_seqs = sum(len(v) for v in self.dataset.values())
            total_frames = total_seqs * self.SEQ_LEN
            
            ws.append([])
            ws.append(['TỔNG CỘNG', '', total_seqs, total_frames, total_seqs])
            
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='F0F9FF', end_color='F0F9FF', fill_type='solid')
            
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            wb.save(filename)
            messagebox.showinfo("Thành công", f"Đã xuất Excel:\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xuất Excel: {e}")
    
    # ===== RECOGNITION WITH GOOGLE TTS =====
    
    def recognize_current(self):
        if not self.model:
            self.lbl_pred.config(text="⚠️")
            self.lbl_conf.config(text="Chưa có mô hình!\nVui lòng train trước.")
            return
        
        # Check buffer is full
        if len(self.frame_buffer) < self.SEQ_LEN:
            self.lbl_pred.config(text="⏳")
            self.lbl_conf.config(text=f"Đang chờ buffer: {len(self.frame_buffer)}/{self.SEQ_LEN}")
            return
        
        seq = np.array(list(self.frame_buffer)).reshape(1, self.SEQ_LEN, 784)
        
        try:
            probs = self.model.predict(seq, verbose=0)[0]
        except Exception as e:
            print(f"❌ Prediction error: {e}")
            self.lbl_pred.config(text="❌")
            self.lbl_conf.config(text="Lỗi dự đoán")
            return
        
        idx = int(np.argmax(probs))
        conf = float(probs[idx])
        
        labels = [lb for lb in self.labels_order if lb in self.dataset and len(self.dataset[lb]) > 0]
        
        if idx >= len(labels):
            self.lbl_pred.config(text="⚠️")
            self.lbl_conf.config(text="Model không khớp dataset")
            return
        
        label = labels[idx]
        
        if conf > 0.6:
            self.lbl_pred.config(text=label)
            self.lbl_conf.config(text=f"Độ tin cậy: {conf*100:.1f}%")
            
            # Google TTS with cooldown (prevent multiple speaks)
            current_time = time.time()
            if label != self.last_spoken or (current_time - self.last_spoken_time > 3.0):
                self.last_spoken = label
                self.last_spoken_time = current_time
                if self.gtts_available:
                    threading.Thread(target=lambda: self._speak_gtts(label), daemon=True).start()
        else:
            self.lbl_pred.config(text="❓")
            self.lbl_conf.config(text=f"Không chắc chắn ({conf*100:.1f}%)")
    
    def _speak_gtts(self, text):
        """Speak using Google TTS - Vietnamese with better error handling"""
        with self.tts_lock:
            if self.tts_playing:
                return
            
            try:
                self.tts_playing = True
                print(f"🔊 Speaking: {text}")
                
                # Generate TTS
                tts = gTTS(text=text, lang='vi', slow=False)
                temp_file = Path("temp_tts.mp3")
                tts.save(str(temp_file))
                
                # Ensure pygame mixer is ready
                if not pygame.mixer.get_init():
                    pygame.mixer.init(frequency=22050, size=-16, channels=2, buffer=512)
                
                # Load and play
                pygame.mixer.music.load(str(temp_file))
                pygame.mixer.music.play()
                
                # Wait for playback to finish
                while pygame.mixer.music.get_busy():
                    pygame.time.Clock().tick(10)
                
                # Cleanup
                pygame.mixer.music.stop()
                pygame.mixer.music.unload()
                
                # Small delay before deleting file
                time.sleep(0.1)
                
                if temp_file.exists():
                    try:
                        temp_file.unlink()
                    except:
                        pass
                
                print(f"✅ TTS completed: {text}")
                
            except Exception as e:
                print(f"❌ TTS Error: {e}")
                import traceback
                traceback.print_exc()
            finally:
                self.tts_playing = False
    
    # ===== SPEECH RECOGNITION =====
    
    def start_listening(self):
        if not sr or not self.recognizer:
            messagebox.showerror("Thiếu thư viện", "Cài: pip install SpeechRecognition pyaudio")
            return
        
        try:
            self.mic = sr.Microphone()
        except Exception as e:
            messagebox.showerror("Lỗi Mic", str(e))
            return
        
        self.stt_is_listening = True
        self.btn_listen.config(state=tk.DISABLED)
        self.btn_stop_listen.config(state=tk.NORMAL)
        self.lbl_mic.config(text="🎙️ Đang nghe...", fg=self.COLOR_SUCCESS)
        self.lock_interface(True)
        
        threading.Thread(target=self._listen_loop, daemon=True).start()
    
    def stop_listening(self):
        self.stt_is_listening = False
        self.btn_listen.config(state=tk.NORMAL)
        self.btn_stop_listen.config(state=tk.DISABLED)
        self.lbl_mic.config(text="🎤 Microphone: Sẵn sàng", fg=self.COLOR_TEXT)
        self.lock_interface(False)
        
        self.playback_running = False
        self.canvas.delete("all")
    
    def _listen_loop(self):
        with self.mic as source:
            self.recognizer.adjust_for_ambient_noise(source, duration=0.6)
        
        while self.stt_is_listening:
            try:
                with self.mic as source:
                    audio = self.recognizer.listen(source, timeout=3, phrase_time_limit=5)
                
                try:
                    text = self.recognizer.recognize_google(audio, language='vi-VN')
                except:
                    text = ""
                
                if text:
                    self.root.after(0, lambda t=text: self.on_speech_text(t))
            except:
                continue
    
    def on_speech_text(self, text):
        self.txt_speech.delete('1.0', tk.END)
        self.txt_speech.insert(tk.END, text)
        
        best, score = self.best_label(text)
        self.lbl_match.config(text=f"Khớp: {best or '—'} ({int(score*100)}%)")
        
        if best and score > 0.5:
            self.play_label(best)
    
    def best_label(self, text):
        labels = [lb for lb in self.labels_order if lb in self.dataset and len(self.dataset[lb]) > 0]
        if not labels:
            return None, 0.0
        
        t = text.strip().lower()
        lst = [s.lower() for s in labels]
        
        m = difflib.get_close_matches(t, lst, n=1, cutoff=0.4)
        if m:
            idx = lst.index(m[0])
            return labels[idx], 1.0
        
        scores = [(lab, difflib.SequenceMatcher(None, t, lab.lower()).ratio()) for lab in labels]
        best = max(scores, key=lambda x: x[1])
        return best[0], float(best[1])
    
    def play_label(self, label):
        """Play animation frames for a label"""
        print(f"\n🔹 Attempting to play label: '{label}'")
        print(f"📁 Frames index: {self.frames_index.get(label, 'NOT FOUND')}")
        
        seq_dirs = self.frames_index.get(label, [])
        if not seq_dirs:
            print(f"❌ No sequences found for label '{label}'")
            messagebox.showwarning("Không có dữ liệu", f"Chưa có animation cho '{label}'")
            return
        
        # Get last sequence
        seq_rel = seq_dirs[-1]
        seq_dir = self.frames_root / seq_rel
        
        print(f"📂 Looking for frames in: {seq_dir}")
        print(f"📂 Absolute path: {seq_dir.absolute()}")
        print(f"📂 Directory exists: {seq_dir.exists()}")
        
        if not seq_dir.exists():
            print(f"❌ Directory does not exist!")
            messagebox.showerror("Lỗi", f"Thư mục không tồn tại:\n{seq_dir}")
            return
        
        frames = sorted(seq_dir.glob("*.jpg"))
        print(f"🖼️ Found {len(frames)} frames")
        
        if not frames:
            print(f"❌ No JPG frames found in {seq_dir}")
            # List all files in directory for debugging
            all_files = list(seq_dir.glob("*"))
            print(f"📄 Files in directory: {[f.name for f in all_files]}")
            messagebox.showerror("Lỗi", f"Không có ảnh trong:\n{seq_dir}")
            return
        
        try:
            self.playback_frames = [Image.open(p) for p in frames]
            print(f"✅ Successfully loaded {len(self.playback_frames)} frames")
        except Exception as e:
            print(f"❌ Error loading frames: {e}")
            messagebox.showerror("Lỗi", f"Không thể tải ảnh: {e}")
            return
        
        self.playback_idx = 0
        self.playback_label = label
        self.playback_running = True
        
        if self.is_running:
            self.stop_camera()
        
        print(f"🎬 Starting playback for '{label}'")
        self.render_playback()
    
    def render_playback(self):
        if not self.playback_running or not self.playback_frames:
            return
        
        try:
            canvas_w = self.canvas.winfo_width() or 800
            canvas_h = self.canvas.winfo_height() or 600
            
            base = Image.new("RGB", (canvas_w, canvas_h), (26, 26, 26))
            
            w, h = int(canvas_w * 0.7), int(canvas_h * 0.7)
            x, y = (canvas_w - w) // 2, (canvas_h - h) // 2
            
            frame_img = self.playback_frames[self.playback_idx % len(self.playback_frames)]
            frame_img = frame_img.resize((w, h), Image.LANCZOS)
            base.paste(frame_img, (x, y))
            
            draw = ImageDraw.Draw(base)
            draw.rectangle((x, y - 45, x + 320, y - 5), fill=(59, 130, 246))
            
            try:
                font = ImageFont.truetype("arial.ttf", 20)
            except:
                font = ImageFont.load_default()
            
            draw.text((x + 15, y - 35), f"🔹 {self.playback_label}", fill=(255, 255, 255), font=font)
            
            photo = ImageTk.PhotoImage(base)
            self.canvas.create_image(0, 0, anchor=tk.NW, image=photo)
            self.canvas.image = photo
            
            self.playback_idx = (self.playback_idx + 1) % len(self.playback_frames)
            self.root.after(60, self.render_playback)
        except Exception as e:
            print(f"Playback error: {e}")
            self.playback_running = False
    
    # ===== UTILS =====
    
    def update_stats(self):
        tot_labels = len([lb for lb in self.dataset if len(self.dataset[lb]) > 0])
        tot_seq = sum(len(v) for v in self.dataset.values())
        
        if self.current_mode == "collect":
            label = self.entry_label.get().strip()
            cur = len(self.dataset.get(label, [])) if label else 0
            self.lbl_collect_stats.config(
                text=f"Sequences: {cur} | Tổng: {tot_seq}"
            )
        
        model_status = f"✅ Sẵn sàng" if self.model else "Chưa train"
        self.lbl_train_stats.config(
            text=f"Dataset: {tot_labels} nhãn, {tot_seq} sequences\nMô hình: {model_status}"
        )
    
    def save_dataset(self):
        with open(self.dataset_pkl, 'wb') as f:
            pickle.dump({
                'dataset': self.dataset,
                'labels_order': self.labels_order
            }, f)
    
    def load_dataset(self):
        if self.dataset_pkl.exists():
            try:
                obj = pickle.load(open(self.dataset_pkl, 'rb'))
                if isinstance(obj, dict) and 'dataset' in obj:
                    self.dataset = obj.get('dataset', {})
                    self.labels_order = obj.get('labels_order', list(self.dataset.keys()))
                else:
                    self.dataset = obj
                    self.labels_order = list(self.dataset.keys())
            except:
                self.dataset = {}
                self.labels_order = []
    
    def save_frames_index(self):
        with open(self.frames_index_path, 'w', encoding='utf-8') as f:
            json.dump(self.frames_index, f, ensure_ascii=False, indent=2)
    
    def load_frames_index(self):
        if self.frames_index_path.exists():
            try:
                self.frames_index = json.load(
                    open(self.frames_index_path, 'r', encoding='utf-8')
                )
                print(f"\n✅ Loaded frames_index.json:")
                for label, seqs in self.frames_index.items():
                    print(f"  '{label}': {len(seqs)} sequences")
            except Exception as e:
                print(f"⚠️ Could not load frames_index.json: {e}")
                self.frames_index = {}
        else:
            print("📝 frames_index.json not found, starting fresh")
            self.frames_index = {}
    
    def load_model(self):
        """Load existing trained model if available"""
        if self.model_path.exists():
            try:
                self.model = keras.models.load_model(str(self.model_path))
                print(f"\n✅ Loaded existing model from {self.model_path}")
                print(f"   Model has {len(self.model.layers)} layers")
                
                # Verify model matches current labels
                expected_outputs = len([lb for lb in self.labels_order if lb in self.dataset and len(self.dataset[lb]) > 0])
                actual_outputs = self.model.output_shape[-1]
                
                if expected_outputs > 0 and expected_outputs != actual_outputs:
                    print(f"⚠️ Warning: Model output mismatch!")
                    print(f"   Model expects {actual_outputs} labels, but dataset has {expected_outputs} labels")
                    print(f"   You may need to retrain the model.")
                else:
                    print(f"   Model matches dataset: {actual_outputs} labels")
                    
            except Exception as e:
                print(f"⚠️ Could not load model: {e}")
                self.model = None
        else:
            print(f"📝 No existing model found at {self.model_path}")
    
    def save_model(self):
        """Save trained model to disk"""
        if self.model:
            try:
                self.model.save(str(self.model_path))
                print(f"\n💾 Model saved to {self.model_path}")
            except Exception as e:
                print(f"❌ Error saving model: {e}")


if __name__ == "__main__":
    print("\n" + "="*80)
    print("🤖 ROBOT PHIÊN DỊCH NGÔN NGỮ KÝ HIỆU 2 CHIỀU v2.1")
    print("Ứng dụng học máy Deep Learning - OPTIMIZED VERSION")
    print("="*80)
    print("\n✨ CẢI TIẾN MỚI:")
    print("  ✅ Password session-based (chỉ nhập 1 lần)")
    print("  🔊 Google TTS tiếng Việt online")
    print("  ⚡ Fix TTS nói nhiều lần")
    print("  🎨 Fix giao diện crash")
    print("  🗑️ Nút xóa CỰC RÕ RÀNG với khung đỏ")
    print("  🚀 Tối ưu performance camera")
    print("  📜 Thanh cuộn dọc cho panel điều khiển")
    print("  📊 Xuất Excel với định dạng đẹp")
    print("  🎯 Nhận diện + TTS tự động")
    print("  🎙️ Giọng nói → Animation")
    print("\n📦 THƯ VIỆN CẦN CÀI:")
    print("  pip install tensorflow opencv-python pillow")
    print("  pip install numpy SpeechRecognition pyaudio")
    print("  pip install gTTS pygame openpyxl")
    print("\n🔒 MẬT KHẨU MẶC ĐỊNH: admin123")
    print("="*80 + "\n")
    
    root = tk.Tk()
    app = UltimateSignLanguageApp(root)
    root.mainloop()